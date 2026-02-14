#!/usr/bin/env python3
"""
KARDS Card Scraper v2
Улучшенная версия с собиранием карт по мере загрузки
"""

import asyncio
import logging
from typing import List, Dict, Optional, Set
from playwright.async_api import async_playwright, Page, Browser
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import sys
import json

# Настройка логирования
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)


class KardsScraper:
    def __init__(self):
        self.url = "https://www.kards.com/ru/decks/collection"
        self.cards: List[Dict] = []
        self.card_titles: Set[str] = set()
        self.browser: Optional[Browser] = None
        self.page: Optional[Page] = None

    async def initialize_browser(self):
        """Инициализация Playwright браузера"""
        logger.info("Инициализация браузера...")
        self.playwright = await async_playwright().start()
        self.browser = await self.playwright.chromium.launch(headless=True)
        self.page = await self.browser.new_page()
        logger.info("Браузер успешно инициализирован")

    async def close_browser(self):
        """Закрытие браузера"""
        if self.page:
            await self.page.close()
        if self.browser:
            await self.browser.close()
        if self.playwright:
            await self.playwright.stop()
        logger.info("Браузер закрыт")

    async def load_page(self):
        """Загрузка страницы и ожидание полной загрузки карт"""
        logger.info(f"Загрузка страницы {self.url}...")
        try:
            await self.page.goto(self.url, wait_until="networkidle", timeout=60000)
            logger.info("Страница загружена")
        except Exception as e:
            logger.error(f"Ошибка при загрузке страницы: {e}")
            raise

    async def extract_visible_cards(self) -> int:
        """Извлечение видимых карт со страницы и добавление в список"""
        try:
            cards_data = await self.page.evaluate("""
                () => {
                    const cards = [];
                    const cardElements = document.querySelectorAll('[class*="Card_card"]');

                    cardElements.forEach(element => {
                        try {
                            const fullText = (element.innerText || element.textContent || '').trim();

                            if (fullText && fullText.length > 0) {
                                cards.push({
                                    text: fullText,
                                    className: element.className
                                });
                            }
                        } catch (e) {
                            // Пропускаем элементы, которые не удалось обработать
                        }
                    });

                    return cards;
                }
            """)

            cards_before = len(self.cards)

            # Обработка извлеченных данных
            for card_data in cards_data:
                lines = [line.strip() for line in card_data['text'].split('\n') if line.strip()]
                if lines:
                    title = lines[0]
                    if title not in self.card_titles:
                        card_info = {
                            'Название': title,
                            'Страна': '',
                            'Тип': self._extract_field(lines, ['Unit', 'Order', 'Countermeasure']),
                            'Стоимость': self._extract_cost(lines),
                            'Редкость': self._extract_field(lines, ['Common', 'Limited', 'Special', 'Elite', 'Rare']),
                            'Атака': '',
                            'Защита': '',
                            'Набор': '',
                            'Описание': card_data['text']
                        }
                        self.cards.append(card_info)
                        self.card_titles.add(title)

            new_cards_count = len(self.cards) - cards_before
            return new_cards_count

        except Exception as e:
            logger.error(f"Ошибка при извлечении видимых карт: {e}")
            return 0

    def _extract_field(self, lines: List[str], keywords: List[str]) -> str:
        """Извлечение поля по ключевым словам"""
        for line in lines:
            for keyword in keywords:
                if keyword in line:
                    return line.strip()
        return ""

    def _extract_cost(self, lines: List[str]) -> str:
        """Извлечение стоимости карты"""
        for line in lines:
            if line and line[0].isdigit():
                return line.strip()
        return ""

    async def load_and_extract_cards(self):
        """Загрузка и извлечение карт по мере их появления"""
        logger.info("Начало загрузки и извлечения карт...")

        load_more_clicks = 0
        max_clicks = 100
        no_new_cards_count = 0

        while load_more_clicks < max_clicks:
            # Извлекаем видимые карты
            new_cards = await self.extract_visible_cards()
            logger.info(f"Извлечено новых карт: {new_cards}, всего карт: {len(self.cards)}")

            if new_cards == 0:
                no_new_cards_count += 1
                if no_new_cards_count > 2:
                    logger.info("Нет новых карт несколько раз подряд")
                    break
            else:
                no_new_cards_count = 0

            # Ищем кнопку "LOAD MORE"
            load_more_button = await self.page.query_selector('button')

            if not load_more_button:
                logger.info("Кнопка 'LOAD MORE' не найдена. Все карты загружены.")
                break

            # Прокручиваем к кнопке
            try:
                await load_more_button.scroll_into_view_if_needed()
                await asyncio.sleep(0.3)

                # Нажимаем кнопку
                await load_more_button.click()
                load_more_clicks += 1
                logger.info(f"Клик на 'LOAD MORE' {load_more_clicks}/{max_clicks}")

                # Ждем загрузки новых карт (уменьшено время)
                await asyncio.sleep(1.5)

            except Exception as e:
                logger.warning(f"Ошибка при клике на LOAD MORE: {e}")
                break

        # Финальное извлечение после прекращения загрузки
        final_new = await self.extract_visible_cards()
        logger.info(f"Финальное извлечение: {final_new} новых карт")
        logger.info(f"Всего извлечено карт: {len(self.cards)}")

    def export_to_excel(self, filename: str = "kards_cards.xlsx"):
        """Экспорт данных в Excel формат"""
        logger.info(f"Экспорт данных в Excel ({filename})...")

        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "KARDS Cards"

            # Заголовки колонок
            headers = ['Название', 'Страна', 'Тип', 'Стоимость', 'Редкость', 'Атака', 'Защита', 'Набор', 'Описание']
            ws.append(headers)

            # Форматирование заголовков
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")

            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

            # Добавление данных карт
            for card in self.cards:
                row = [
                    card.get('Название', ''),
                    card.get('Страна', ''),
                    card.get('Тип', ''),
                    card.get('Стоимость', ''),
                    card.get('Редкость', ''),
                    card.get('Атака', ''),
                    card.get('Защита', ''),
                    card.get('Набор', ''),
                    card.get('Описание', '')
                ]
                ws.append(row)

            # Автоширина колонок
            column_widths = [30, 15, 15, 15, 15, 10, 10, 20, 50]
            for i, width in enumerate(column_widths, 1):
                ws.column_dimensions[get_column_letter(i)].width = width

            # Заморозка первой строки
            ws.freeze_panes = "A2"

            # Добавление фильтров
            ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(self.cards) + 1}"

            # Сохранение файла
            wb.save(filename)
            logger.info(f"Excel файл успешно создан: {filename}")
            logger.info(f"Всего карт в файле: {len(self.cards)}")

        except Exception as e:
            logger.error(f"Ошибка при экспорте в Excel: {e}")
            raise

    async def run(self):
        """Основной метод запуска скрейпера"""
        try:
            logger.info("Запуск KARDS Card Scraper v2...")
            logger.info(f"Целевая страница: {self.url}")

            await self.initialize_browser()
            await self.load_page()
            await self.load_and_extract_cards()

            if len(self.cards) > 0:
                self.export_to_excel()
            else:
                logger.warning("Не найдено ни одной карты")

            logger.info("Скрейпер успешно завершил работу!")
            return True

        except Exception as e:
            logger.error(f"Критическая ошибка: {e}")
            import traceback
            logger.error(traceback.format_exc())
            return False

        finally:
            await self.close_browser()


async def main():
    """Точка входа программы"""
    scraper = KardsScraper()
    success = await scraper.run()
    sys.exit(0 if success else 1)


if __name__ == "__main__":
    asyncio.run(main())
