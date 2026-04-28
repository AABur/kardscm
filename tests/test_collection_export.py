"""Tests for collection export behavior."""

from __future__ import annotations

import json
from pathlib import Path

import pytest
from openpyxl import Workbook

from kardscm.commands import export_collection
from kardscm.config import LANGUAGE_RU
from kardscm.export import add_deck_sheet, export_deck_to_json


def test_export_requires_data(tmp_path: Path) -> None:
    db_path = tmp_path / "collection.db"
    output_path = tmp_path / "out.csv"

    with pytest.raises(SystemExit, match="Run 'kards sync' first"):
        export_collection("csv", str(output_path), db_path=str(db_path))


@pytest.fixture()
def deck_meta():
    return {
        "name": "Test Deck",
        "major_power": "soviet",
        "ally": "usa",
        "hq": "STALINGRAD",
        "deck_code": "%%CODE",
    }


@pytest.fixture()
def deck_cards():
    return [
        {
            "faction": "Soviet",
            "title": json.dumps({"en-EN": "16th Rifle Regiment", "ru-RU": "16-й СТРЕЛКОВЫЙ ПОЛК"}),
            "type": "infantry",
            "rarity": "Standard",
            "attributes": json.dumps(["guard"]),
            "set": "Base",
            "kredits": 1,
            "attack": 1,
            "defense": 2,
            "text": json.dumps({"en-EN": "Test", "ru-RU": "Тест"}),
            "deck_quantity": 2,
            "deck_cost": 1,
        },
    ]


def test_add_deck_sheet(deck_meta, deck_cards) -> None:
    wb = Workbook()
    lang = LANGUAGE_RU
    add_deck_sheet(
        wb,
        deck_meta,
        deck_cards,
        lang.deck_headers,
        lang.deck_metadata_labels,
        lang.nation_display_names,
        lang,
    )

    assert "Test Deck" in wb.sheetnames
    ws = wb["Test Deck"]
    assert ws.cell(row=1, column=1).value == "Название"
    assert ws.cell(row=1, column=2).value == "Test Deck"
    assert ws.cell(row=2, column=2).value == "Советский Союз"
    assert ws.cell(row=3, column=2).value == "США"
    assert ws.cell(row=7, column=1).value == "Карта"
    assert ws.cell(row=8, column=1).value == "Советские"
    assert ws.cell(row=9, column=1).value == "16-й СТРЕЛКОВЫЙ ПОЛК"


def test_export_deck_to_json(tmp_path: Path, deck_meta, deck_cards) -> None:
    output = tmp_path / "deck.json"
    lang = LANGUAGE_RU
    export_deck_to_json(deck_meta, deck_cards, str(output), lang)

    data = json.loads(output.read_text(encoding="utf-8"))
    assert data["deck"]["name"] == "Test Deck"
    assert "deck_code" not in data["deck"]
    assert len(data["cards"]) == 1
    assert data["cards"][0]["deck_quantity"] == 2
