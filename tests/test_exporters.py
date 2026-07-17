"""Tests for kardscm.export.exporters."""

from __future__ import annotations

import json

from openpyxl import load_workbook

from kardscm.export import (
    build_collection_headers,
    export_to_json,
    export_to_xlsx,
)
from kardscm.export.collection import COLLECTION_TABLE_FIELDS
from kardscm.locales import LANGUAGE_EN, LANGUAGE_RU


def _raw_cards() -> list[dict]:
    """Sample RAW database rows for the JSON export path.

    Title/text/can_create are JSON strings, abilities are binary columns,
    and the remaining fields use the camelCase API names as stored.
    """
    return [
        {
            "cardId": "card-usa-1",
            "importId": "imp-1",
            "imageUrl": "https://example.com/a.png",
            "thumbUrl": "https://example.com/a-thumb.png",
            "faction": "USA",
            "type": "infantry",
            "rarity": "Standard",
            "set": "Base",
            "title": json.dumps({"en-EN": "Alpha", "ru-RU": "Альфа"}),
            "text": json.dumps({"en-EN": "Test card", "ru-RU": "Тестовая карта"}),
            "kredits": 1,
            "attack": 1,
            "defense": 2,
            "ability_guard": 1,
            "ability_blitz": 0,
            "extra_ability_naval": 1,
            "extra_ability_pincer": 0,
            "operationCost": 3,
            "reserved": 0,
            "image": "a.png",
            "can_create": None,
            "exile": None,
            "quantity": 2,
        },
        {
            "cardId": "card-de-1",
            "importId": "imp-2",
            "imageUrl": "https://example.com/b.png",
            "thumbUrl": "https://example.com/b-thumb.png",
            "faction": "Germany",
            "type": "tank",
            "rarity": "Limited",
            "set": "Base",
            "title": json.dumps({"en-EN": "Panzer IV", "ru-RU": "Панцер IV"}),
            "text": json.dumps({"en-EN": "Tank card", "ru-RU": "Танк"}),
            "kredits": 4,
            "attack": 3,
            "defense": 4,
            "ability_guard": 0,
            "ability_blitz": 1,
            "extra_ability_naval": 0,
            "extra_ability_pincer": 0,
            "operationCost": None,
            "reserved": 0,
            "image": "b.png",
            "can_create": None,
            "exile": None,
            "quantity": 1,
        },
    ]


def _translated_cards() -> list[dict]:
    """Sample translated card dicts (12 web-table fields) for the XLSX path."""
    return [
        {
            "faction": "USA",
            "title": "Alpha",
            "type": "Infantry",
            "rarity": "Standard",
            "abilities": "Guard",
            "extra_abilities": "Naval",
            "set": "Base",
            "quantity": 2,
            "kredits": 1,
            "operationCost": 3,
            "attack": 1,
            "defense": 2,
            "text": "Test card",
        },
        {
            "faction": "Germany",
            "title": "Panzer IV",
            "type": "Tank",
            "rarity": "Limited",
            "abilities": "Blitz",
            "extra_abilities": "",
            "set": "Base",
            "quantity": 1,
            "kredits": 4,
            "operationCost": None,
            "attack": 3,
            "defense": 4,
            "text": "Tank card",
        },
    ]


class TestBuildCollectionHeaders:
    def test_matches_web_table_order(self) -> None:
        assert build_collection_headers(LANGUAGE_EN) == [
            "Nation",
            "Name",
            "Type",
            "Rarity",
            "Abilities",
            "Extra abilities",
            "Set",
            "Quantity",
            "Credits",
            "Cost",
            "Attack",
            "Defense",
        ]

    def test_xlsx_headers_match_web_table_order(self) -> None:
        """The XLSX header source must mirror the rendered web table headers."""
        headers = build_collection_headers(LANGUAGE_EN)
        assert headers[0] == "Nation"
        assert headers[5] == "Extra abilities"
        assert headers[9] == "Cost"
        assert "Description" not in headers
        assert len(headers) == 12

    def test_localized_headers(self) -> None:
        headers = build_collection_headers(LANGUAGE_RU)
        assert len(headers) == 12
        assert "Description" not in headers


class TestExportToXlsx:
    def test_creates_file(self, tmp_path):
        out = tmp_path / "cards.xlsx"
        cards = _translated_cards()
        headers = build_collection_headers(LANGUAGE_EN)

        export_to_xlsx(cards, str(out), headers)

        assert out.exists()
        wb = load_workbook(str(out))
        ws = wb.active
        header_values = [cell.value for cell in ws[1]]
        assert header_values == headers
        assert ws.cell(row=2, column=2).value == "Alpha"

    def test_header_row_has_12_columns(self, tmp_path):
        out = tmp_path / "cards.xlsx"
        headers = build_collection_headers(LANGUAGE_EN)
        export_to_xlsx(_translated_cards(), str(out), headers)

        wb = load_workbook(str(out))
        ws = wb.active
        header_values = [cell.value for cell in ws[1]]
        assert header_values == headers
        assert "Extra abilities" in header_values
        assert "Cost" in header_values
        assert "Description" not in header_values
        assert len(COLLECTION_TABLE_FIELDS) == 12

    def test_extra_abilities_column_value(self, tmp_path):
        out = tmp_path / "cards.xlsx"
        headers = build_collection_headers(LANGUAGE_EN)
        export_to_xlsx(_translated_cards(), str(out), headers)

        wb = load_workbook(str(out))
        ws = wb.active
        extra_col = headers.index("Extra abilities") + 1
        assert ws.cell(row=2, column=extra_col).value == "Naval"

    def test_cost_column_value(self, tmp_path):
        out = tmp_path / "cards.xlsx"
        headers = build_collection_headers(LANGUAGE_EN)
        export_to_xlsx(_translated_cards(), str(out), headers)

        wb = load_workbook(str(out))
        ws = wb.active
        cost_col = headers.index("Cost") + 1
        assert ws.cell(row=2, column=cost_col).value == 3

    def test_quantity_column_index_shifted(self, tmp_path):
        out = tmp_path / "cards.xlsx"
        headers = build_collection_headers(LANGUAGE_EN)
        export_to_xlsx(_translated_cards(), str(out), headers)

        wb = load_workbook(str(out))
        ws = wb.active
        qty_col = headers.index("Quantity") + 1
        # Quantity is now the 8th column (after the Extra abilities insertion).
        assert qty_col == 8
        assert ws.cell(row=2, column=qty_col).value == 2

    def test_correct_row_count(self, tmp_path):
        out = tmp_path / "cards.xlsx"
        export_to_xlsx(_translated_cards(), str(out), build_collection_headers(LANGUAGE_EN))

        wb = load_workbook(str(out))
        ws = wb.active
        assert ws.max_row == 3

    def test_freeze_panes(self, tmp_path):
        out = tmp_path / "cards.xlsx"
        export_to_xlsx(_translated_cards(), str(out), build_collection_headers(LANGUAGE_EN))

        wb = load_workbook(str(out))
        ws = wb.active
        assert ws.freeze_panes == "A2"


class TestExportToJson:
    def test_metadata_has_only_total_cards(self, tmp_path):
        out = tmp_path / "cards.json"
        export_to_json(_raw_cards(), str(out))

        data = json.loads(out.read_text(encoding="utf-8"))
        assert data["metadata"] == {"total_cards": 2}
        assert "language" not in data["metadata"]
        assert "language_name" not in data["metadata"]

    def test_title_is_dict(self, tmp_path):
        out = tmp_path / "cards.json"
        export_to_json(_raw_cards(), str(out))

        data = json.loads(out.read_text(encoding="utf-8"))
        card = data["cards"][0]
        assert isinstance(card["title"], dict)
        assert card["title"]["en-EN"] == "Alpha"
        assert card["title"]["ru-RU"] == "Альфа"

    def test_attributes_and_extra_abilities_are_lists(self, tmp_path):
        out = tmp_path / "cards.json"
        export_to_json(_raw_cards(), str(out))

        data = json.loads(out.read_text(encoding="utf-8"))
        card = data["cards"][0]
        assert card["attributes"] == ["guard"]
        assert card["extra_abilities"] == ["naval"]
        assert isinstance(card["attributes"], list)
        assert isinstance(card["extra_abilities"], list)

    def test_raw_codes_not_localized(self, tmp_path):
        out = tmp_path / "cards.json"
        export_to_json(_raw_cards(), str(out))

        data = json.loads(out.read_text(encoding="utf-8"))
        card = data["cards"][0]
        assert card["faction"] == "USA"
        assert card["type"] == "infantry"
        assert card["rarity"] == "Standard"
        assert card["set"] == "Base"

    def test_card_includes_all_api_fields(self, tmp_path):
        out = tmp_path / "cards.json"
        export_to_json(_raw_cards(), str(out))

        data = json.loads(out.read_text(encoding="utf-8"))
        card = data["cards"][0]
        for key in (
            "cardId",
            "importId",
            "imageUrl",
            "thumbUrl",
            "kredits",
            "attack",
            "defense",
            "operationCost",
            "quantity",
            "reserved",
            "image",
            "can_create",
            "exile",
        ):
            assert key in card
        assert card["cardId"] == "card-usa-1"
        assert card["operationCost"] == 3
        assert card["quantity"] == 2

    def test_lang_invariance(self, tmp_path):
        """The raw JSON payload must be identical regardless of --lang."""
        raw = _raw_cards()
        out_en = tmp_path / "en.json"
        out_ru = tmp_path / "ru.json"
        export_to_json(raw, str(out_en))
        export_to_json(raw, str(out_ru))

        data_en = json.loads(out_en.read_text(encoding="utf-8"))
        data_ru = json.loads(out_ru.read_text(encoding="utf-8"))
        assert data_en["cards"] == data_ru["cards"]

    def test_creates_file(self, tmp_path):
        out = tmp_path / "cards.json"
        export_to_json(_raw_cards(), str(out))
        assert out.exists()


class TestExportEmptyCards:
    def test_xlsx_empty(self, tmp_path):
        out = tmp_path / "empty.xlsx"
        export_to_xlsx([], str(out), build_collection_headers(LANGUAGE_EN))
        assert out.exists()
        wb = load_workbook(str(out))
        ws = wb.active
        assert ws.max_row == 1

    def test_json_empty(self, tmp_path):
        out = tmp_path / "empty.json"
        export_to_json([], str(out))
        data = json.loads(out.read_text(encoding="utf-8"))
        assert data["metadata"]["total_cards"] == 0
        assert data["cards"] == []
