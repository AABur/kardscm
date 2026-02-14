"""Business logic for CLI commands."""

from __future__ import annotations

import logging
import sqlite3
from datetime import UTC, datetime
from pathlib import Path

from openpyxl import load_workbook

from kards.constants import (
    DECK_NATION_TO_DB,
    DEFAULT_DB_PATH,
    LANGUAGE_CODE,
    LANGUAGE_NAME,
    RUSSIAN_HEADERS,
)
from kards.export import (
    add_deck_sheet,
    export_deck_to_json,
    export_to_csv,
    export_to_json,
    export_to_xlsx,
)
from kards.helpers import parse_int
from kards.importing import parse_deck_file
from kards.scraping import scrape_cards
from kards.storage import (
    fetch_all_decks,
    fetch_cards,
    fetch_deck_cards,
    find_card_id_by_nation_name,
    find_deck_by_name,
    get_connection,
    initialize_schema,
    insert_deck,
    insert_deck_cards,
    set_metadata,
    update_quantity_by_nation_name,
    upsert_cards,
)

logger = logging.getLogger(__name__)


def _utc_timestamp() -> str:
    return datetime.now(UTC).replace(tzinfo=None).isoformat(timespec="seconds")


def validate_file(path: str, expected_ext: str, must_exist: bool = False) -> Path:
    """Validate file extension and existence.

    Args:
        path: File path string.
        expected_ext: Expected extension (e.g. ".xlsx").
        must_exist: If True, file must exist on disk.

    Returns:
        Resolved Path object.

    Raises:
        SystemExit: If validation fails.
    """
    p = Path(path)
    if p.suffix != expected_ext:
        raise SystemExit(f"Expected {expected_ext} file, got: {p.suffix or '(no extension)'}")
    if must_exist and not p.exists():
        raise SystemExit(f"File not found: {path}")
    return p


def _read_xlsx_quantities(filename: str) -> list[tuple[str, str, int | None]]:
    """Read nation, name, quantity from XLSX file.

    Args:
        filename: Path to XLSX file.

    Returns:
        List of (nation, name, quantity) tuples.

    Raises:
        FileNotFoundError: If file doesn't exist.
        ValueError: If required columns missing.
    """
    if not Path(filename).exists():
        raise FileNotFoundError(f"File not found: {filename}")

    wb = load_workbook(filename)
    ws = wb.active
    if not ws:
        raise ValueError("No active worksheet found")

    header_map = {"Нация": "nation", "Название": "name", "Количество": "quantity"}
    headers = {
        header_map[cell.value]: col_idx
        for col_idx, cell in enumerate(ws[1], 1)
        if cell.value in header_map
    }
    if len(headers) != len(header_map):
        raise ValueError("Missing required columns: Нация, Название, Количество")

    results = []
    for row in ws.iter_rows(min_row=2, values_only=False):
        nation_cell = row[headers["nation"] - 1]
        name_cell = row[headers["name"] - 1]
        qty_cell = row[headers["quantity"] - 1]

        nation = nation_cell.value
        name = name_cell.value
        qty_val = qty_cell.value

        if not nation or not name:
            continue

        qty = parse_int(qty_val)

        results.append((str(nation).strip(), str(name).strip(), qty))

    return results


async def sync_collection(db_path: str = DEFAULT_DB_PATH) -> None:
    """Synchronize local SQLite storage with the website.

    Args:
        db_path: SQLite database path.
    """
    logger.info("Starting sync from website...")
    cards = await scrape_cards()

    with get_connection(db_path) as conn:
        initialize_schema(conn)
        upsert_cards(conn, cards)
        set_metadata(conn, "last_sync", _utc_timestamp())

    logger.info("Sync completed. Stored %s cards.", len(cards))


def export_collection(
    export_format: str,
    filename: str,
    db_path: str = DEFAULT_DB_PATH,
) -> None:
    """Export cards from SQLite to the selected format.

    Args:
        export_format: Export format (csv, json, xlsx).
        filename: Output file path.
        db_path: SQLite database path.
    """
    with get_connection(db_path) as conn:
        initialize_schema(conn)
        cards = fetch_cards(conn)

    if not cards:
        raise SystemExit("No cards in database. Run 'kards sync' first.")

    if export_format == "xlsx":
        export_to_xlsx(cards, filename, RUSSIAN_HEADERS)
    elif export_format == "csv":
        export_to_csv(cards, filename, RUSSIAN_HEADERS)
    elif export_format == "json":
        export_to_json(cards, filename, LANGUAGE_CODE, LANGUAGE_NAME)
    else:
        msg = f"Unsupported format: {export_format}"
        raise ValueError(msg)

    logger.info("Export completed: %s", filename)


def update_collection(filename: str, db_path: str = DEFAULT_DB_PATH) -> None:
    """Update card quantities from XLSX file.

    Args:
        filename: XLSX file path.
        db_path: SQLite database path.
    """
    logger.info("Starting update from file: %s", filename)

    try:
        updates = _read_xlsx_quantities(filename)
    except (FileNotFoundError, ValueError) as e:
        raise SystemExit(f"Failed to read file: {e}") from e

    if not updates:
        logger.warning("No valid entries found in file")
        return

    with get_connection(db_path) as conn:
        initialize_schema(conn)
        updated_count, not_found = update_quantity_by_nation_name(conn, updates)

    logger.info(
        "Update completed: %d cards updated, %d not found",
        updated_count,
        len(not_found),
    )

    if not_found:
        for key in not_found:
            logger.warning("Card not found: %s", key)


def import_deck(filename: str, db_path: str = DEFAULT_DB_PATH) -> None:
    """Import a deck from TXT file into the database.

    Args:
        filename: Path to deck TXT file.
        db_path: SQLite database path.
    """
    logger.info("Importing deck from file: %s", filename)

    try:
        deck = parse_deck_file(filename)
    except (FileNotFoundError, ValueError) as e:
        raise SystemExit(f"Failed to parse deck file: {e}") from e

    with get_connection(db_path) as conn:
        initialize_schema(conn)

        existing = find_deck_by_name(conn, deck["name"])
        if existing:
            raise SystemExit(f"Deck '{deck['name']}' already exists (id={existing['deck_id']})")

        not_found = []
        for card in deck["cards"]:
            db_nation = DECK_NATION_TO_DB.get(card["nation"], card["nation"])
            card_id = find_card_id_by_nation_name(conn, db_nation, card["name"])
            if card_id is None:
                not_found.append(f"{db_nation} / {card['name']}")

        if not_found:
            lines = "\n".join(f"  - {entry}" for entry in not_found)
            raise SystemExit(f"Cards not found in collection:\n{lines}")

        deck_id = insert_deck(conn, deck)
        insert_deck_cards(conn, deck_id, deck["cards"], DECK_NATION_TO_DB)
        conn.commit()

    logger.info("Deck '%s' imported (%d cards)", deck["name"], len(deck["cards"]))


def _select_deck(conn: sqlite3.Connection) -> dict:
    """Interactively select a deck from the database.

    Args:
        conn: SQLite connection instance.

    Returns:
        Selected deck metadata dict.
    """
    decks = fetch_all_decks(conn)
    if not decks:
        raise SystemExit("No decks in database. Run 'kards deck import' first.")

    print("Available decks:")
    for i, deck in enumerate(decks, 1):
        print(f"  {i}. {deck['name']}")

    try:
        choice = int(input("Enter deck number: "))
    except (ValueError, EOFError) as e:
        raise SystemExit("Invalid input") from e

    if choice < 1 or choice > len(decks):
        raise SystemExit(f"Invalid choice: {choice}")

    return decks[choice - 1]


def export_deck(
    fmt: str | None,
    filename: str,
    db_path: str = DEFAULT_DB_PATH,
) -> None:
    """Export a deck to XLSX sheet or JSON file.

    Args:
        fmt: Export format ('json' or None for xlsx).
        filename: Output file path.
        db_path: SQLite database path.
    """
    with get_connection(db_path) as conn:
        initialize_schema(conn)
        deck_meta = _select_deck(conn)
        deck_cards = fetch_deck_cards(conn, deck_meta["deck_id"])

    if not deck_cards:
        raise SystemExit("Deck has no cards")

    if fmt == "json":
        export_deck_to_json(deck_meta, deck_cards, filename)
    else:
        wb = load_workbook(filename)
        add_deck_sheet(wb, deck_meta, deck_cards)
        wb.save(filename)

    logger.info("Deck exported: %s", filename)
