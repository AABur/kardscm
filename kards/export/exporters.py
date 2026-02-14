"""Export helpers for card collections."""

from __future__ import annotations

import csv
import json
import logging
from collections.abc import Iterable

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from kards.constants import EXPORT_FIELD_NAMES

logger = logging.getLogger(__name__)


def export_to_xlsx(
    cards: list[dict[str, str]],
    filename: str,
    headers: Iterable[str] = EXPORT_FIELD_NAMES,
) -> None:
    """Export cards to Excel format with formatting.

    Args:
        cards: List of card dictionaries.
        filename: Output filename.
        headers: Headers to use for the file.
    """
    logger.info("Exporting %s cards to Excel: %s", len(cards), filename)

    wb = Workbook()
    ws = wb.active
    if not ws:
        msg = "Failed to create worksheet"
        raise RuntimeError(msg)
    ws.title = "Cards"

    header_row = list(headers)
    ws.append(header_row)

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for card in cards:
        row = [card.get(field, "") for field in EXPORT_FIELD_NAMES]
        ws.append(row)

    column_widths = [15, 35, 18, 15, 12, 20, 10, 10, 8, 8, 60]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(header_row))}{len(cards) + 1}"

    wb.save(filename)
    logger.info("Excel file created: %s (%s cards)", filename, len(cards))


def export_to_csv(
    cards: list[dict[str, str]],
    filename: str,
    headers: Iterable[str] = EXPORT_FIELD_NAMES,
) -> None:
    """Export cards to CSV format.

    Uses UTF-8 with BOM for Excel compatibility on Windows.

    Args:
        cards: List of card dictionaries.
        filename: Output filename.
        headers: Headers to use for the file.
    """
    logger.info("Exporting %s cards to CSV: %s", len(cards), filename)

    header_row = list(headers)

    with open(filename, "w", newline="", encoding="utf-8-sig") as csvfile:
        writer = csv.writer(csvfile)
        writer.writerow(header_row)

        for card in cards:
            row = [card.get(field, "") for field in EXPORT_FIELD_NAMES]
            writer.writerow(row)

    logger.info("CSV file created: %s (%s cards)", filename, len(cards))


def export_to_json(
    cards: list[dict[str, str]],
    filename: str,
    language: str,
    language_name: str,
) -> None:
    """Export cards to JSON format with metadata.

    Args:
        cards: List of card dictionaries.
        filename: Output filename.
        language: Language code for metadata.
        language_name: Language name for metadata.
    """
    logger.info("Exporting %s cards to JSON: %s", len(cards), filename)

    output_data = {
        "metadata": {
            "language": language,
            "language_name": language_name,
            "total_cards": len(cards),
        },
        "cards": cards,
    }

    with open(filename, "w", encoding="utf-8") as jsonfile:
        json.dump(output_data, jsonfile, ensure_ascii=False, indent=2)

    logger.info("JSON file created: %s (%s cards)", filename, len(cards))
