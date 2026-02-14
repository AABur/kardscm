#!/usr/bin/env python3
"""
KARDS Card Scraper v3 - API version
Перехватывает API запросы для получения данных карт
"""

import asyncio
import logging
import json
from typing import List, Dict, Optional, Set
from playwright.async_api import async_playwright, Page, Browser, Route
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import sys

# Настройка логирования
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)


class KardsAPIScraper:
    def __init__(self):
        self.url = "https://www.kards.com/ru/decks/collection"
        self.cards: List[Dict] = []
        self.card_titles: Set[str] = set()
        self.browser: Optional[Browser] = None
        self.page: Optional[Page] = None
        self.api_responses: List[str] = []

    async def initialize_browser(self):
        """Инициализация Playwright браузера"""
        logger.info("Инициализация браузера...")
        self.playwright = await async_playwright().start()
        self.browser = await self.playwright.chromium.launch(headless=True)
        self.page = await self.browser.new_page()

        # Регистрируем обработчик для всех сетевых ответов
        async def handle_route(route: Route):
            request = route.request
            url = request.url

            # Интересуют нас запросы к API
            if 'api' in url or 'graphql' in url or 'collection' in url:
                try:
                    response = await route.fetch()
                    status = response.status

                    logger.info(f"API запрос: {url} [статус: {status}]")

                    # Пытаемся получить текст ответа
                    if response.ok:
                        try:
                            text = await response.text()
                            if text and len(text) > 0:
                                self.api_responses.append(text)
                                logger.info(f"Получен ответ размером: {len(text)} символов")
                        except Exception as e:
                            logger.warning(f"Не удалось получить текст ответа: {e}")

                    await route.continue_()
                except Exception as e:
                    logger.warning(f"Ошибка при обработке маршрута: {e}")
                    await route.continue_()
            else:
                await route.continue_()

        # Устанавливаем обработчик маршрутов
        await self.page.route("**/*", handle_route)

        logger.info("Браузер успешно инициализирован")

    async def close_browser(self):
        """Закрытие браузера"""
        if self.page:
            await self.page.close()
        if self.browser:
            await self.browser.close()
        if self.playwright:
            await self.playwright.stop()
        logger.info("Браузер закрыт")

    async def load_page(self):
        """Загрузка страницы и ожидание полной загрузки карт"""
        logger.info(f"Загрузка страницы {self.url}...")
        try:
            await self.page.goto(self.url, wait_until="networkidle", timeout=60000)
            logger.info("Страница загружена")
        except Exception as e:
            logger.error(f"Ошибка при загрузке страницы: {e}")
            raise

    async def load_all_cards(self):
        """Загрузка всех карт путем нажатия на кнопку 'LOAD MORE'"""
        logger.info("Загрузка всех карт путем нажатия 'LOAD MORE'...")

        load_more_clicks = 0
        max_clicks = 50

        while load_more_clicks < max_clicks:
            # Ищем кнопку "LOAD MORE"
            buttons = await self.page.query_selector_all('button')

            if not buttons:
                logger.info("Кнопки не найдены")
                break

            load_more_button = None
            for button in buttons:
                text = await button.inner_text()
                if "LOAD MORE" in text.upper():
                    load_more_button = button
                    break

            if not load_more_button:
                logger.info("Кнопка 'LOAD MORE' не найдена. Все карты загружены.")
                break

            # Прокручиваем к кнопке
            try:
                await load_more_button.scroll_into_view_if_needed()
                await asyncio.sleep(0.5)

                # Нажимаем кнопку
                await load_more_button.click()
                load_more_clicks += 1
                logger.info(f"Клик на 'LOAD MORE' {load_more_clicks}/{max_clicks}")

                # Ждем загрузки новых карт
                await asyncio.sleep(2)

            except Exception as e:
                logger.warning(f"Ошибка при клике на LOAD MORE: {e}")
                break

        logger.info(f"Загрузка завершена. Кликов: {load_more_clicks}")

    def parse_api_responses(self):
        """Парсинг перехваченных API ответов для извлечения карт"""
        logger.info(f"Парсинг {len(self.api_responses)} API ответов...")

        for i, response_text in enumerate(self.api_responses):
            try:
                # Пытаемся распарсить как JSON
                data = json.loads(response_text)

                # Ищем данные карт в различных структурах
                self._extract_cards_from_json(data)

            except json.JSONDecodeError:
                logger.debug(f"Ответ {i} не является JSON")
            except Exception as e:
                logger.warning(f"Ошибка при парсинге ответа {i}: {e}")

        logger.info(f"Всего извлечено карт: {len(self.cards)}")

    def _extract_cards_from_json(self, data: any, depth: int = 0):
        """Рекурсивно ищет данные карт в JSON структуре"""
        if depth > 10:  # Ограничиваем глубину рекурсии
            return

        if isinstance(data, dict):
            # Ищем массивы или объекты с похожими на карты данными
            for key, value in data.items():
                if key.lower() in ['cards', 'card', 'items', 'data', 'result', 'results']:
                    if isinstance(value, list):
                        for item in value:
                            self._process_card_item(item)
                    elif isinstance(value, dict):
                        self._process_card_item(value)

                # Продолжаем рекурсию
                self._extract_cards_from_json(value, depth + 1)

        elif isinstance(data, list):
            # Проверяем, это массив карт?
            if len(data) > 0 and isinstance(data[0], dict):
                # Ищем объекты с полями карты
                for item in data:
                    if self._looks_like_card(item):
                        self._process_card_item(item)

    def _looks_like_card(self, obj: dict) -> bool:
        """Проверка, похож ли объект на карту"""
        card_like_keys = ['name', 'title', 'card_name', 'cardName', 'id', 'type', 'cost', 'rarity']
        obj_keys = obj.keys() if isinstance(obj, dict) else []
        return sum(1 for key in card_like_keys if key in obj_keys) >= 2

    def _process_card_item(self, item: dict):
        """Обработка объекта карты"""
        if not isinstance(item, dict):
            return

        # Пытаемся выделить различные поля названия карты
        title = (
            item.get('name') or
            item.get('title') or
            item.get('card_name') or
            item.get('cardName') or
            item.get('card_title') or
            item.get('displayName') or
            ''
        ).strip()

        if title and title not in self.card_titles:
            card_info = {
                'Название': title,
                'Страна': str(item.get('country') or item.get('nation') or item.get('faction') or ''),
                'Тип': str(item.get('type') or item.get('card_type') or ''),
                'Стоимость': str(item.get('cost') or item.get('credit') or item.get('kredits') or ''),
                'Редкость': str(item.get('rarity') or item.get('rarity_level') or ''),
                'Атака': str(item.get('attack') or item.get('atk') or item.get('power') or ''),
                'Защита': str(item.get('defense') or item.get('def') or item.get('health') or ''),
                'Набор': str(item.get('set') or item.get('collection') or item.get('set_name') or ''),
                'Описание': str(item.get('description') or item.get('text') or item.get('effect') or '')
            }

            self.cards.append(card_info)
            self.card_titles.add(title)

    def export_to_excel(self, filename: str = "kards_cards.xlsx"):
        """Экспорт данных в Excel формат"""
        logger.info(f"Экспорт данных в Excel ({filename})...")

        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "KARDS Cards"

            # Заголовки колонок
            headers = ['Название', 'Страна', 'Тип', 'Стоимость', 'Редкость', 'Атака', 'Защита', 'Набор', 'Описание']
            ws.append(headers)

            # Форматирование заголовков
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")

            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

            # Добавление данных карт
            for card in self.cards:
                row = [
                    card.get('Название', ''),
                    card.get('Страна', ''),
                    card.get('Тип', ''),
                    card.get('Стоимость', ''),
                    card.get('Редкость', ''),
                    card.get('Атака', ''),
                    card.get('Защита', ''),
                    card.get('Набор', ''),
                    card.get('Описание', '')
                ]
                ws.append(row)

            # Автоширина колонок
            column_widths = [30, 15, 15, 15, 15, 10, 10, 20, 50]
            for i, width in enumerate(column_widths, 1):
                ws.column_dimensions[get_column_letter(i)].width = width

            # Заморозка первой строки
            ws.freeze_panes = "A2"

            # Добавление фильтров
            ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(self.cards) + 1}"

            # Сохранение файла
            wb.save(filename)
            logger.info(f"Excel файл успешно создан: {filename}")
            logger.info(f"Всего карт в файле: {len(self.cards)}")

        except Exception as e:
            logger.error(f"Ошибка при экспорте в Excel: {e}")
            raise

    async def run(self):
        """Основной метод запуска скрейпера"""
        try:
            logger.info("Запуск KARDS Card Scraper v3 (API версия)...")
            logger.info(f"Целевая страница: {self.url}")

            await self.initialize_browser()
            await self.load_page()
            await self.load_all_cards()

            # Сохраняем API ответы для анализа
            logger.info(f"Получено {len(self.api_responses)} API ответов")
            if self.api_responses:
                with open('api_responses.txt', 'w', encoding='utf-8') as f:
                    for i, response in enumerate(self.api_responses):
                        f.write(f"\n{'='*50}\n")
                        f.write(f"Response {i+1} (length: {len(response)})\n")
                        f.write(f"{'='*50}\n")
                        f.write(response[:2000])  # Первые 2000 символов каждого ответа
                logger.info("API ответы сохранены в 'api_responses.txt'")

            # Парсим API ответы
            self.parse_api_responses()

            if len(self.cards) > 0:
                self.export_to_excel()
            else:
                logger.warning("Не найдено ни одной карты")

            logger.info("Скрейпер успешно завершил работу!")
            return True

        except Exception as e:
            logger.error(f"Критическая ошибка: {e}")
            import traceback
            logger.error(traceback.format_exc())
            return False

        finally:
            await self.close_browser()


async def main():
    """Точка входа программы"""
    scraper = KardsAPIScraper()
    success = await scraper.run()
    sys.exit(0 if success else 1)


if __name__ == "__main__":
    asyncio.run(main())
