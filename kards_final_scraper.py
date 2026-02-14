#!/usr/bin/env python3
"""KARDS Card Scraper - Multi-language & Multi-format Export."""

import argparse
import asyncio
import csv
import json
import logging
import re
import sys
import traceback
from typing import Any

import httpx
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from playwright.async_api import Browser, Page, async_playwright

# Language and format support
SUPPORTED_LANGUAGES: dict[str, str] = {
    "de": "German",
    "en": "English",
    "es": "Spanish",
    "fr": "French",
    "it": "Italian",
    "ja": "Japanese",
    "ko": "Korean",
    "pl": "Polish",
    "pt": "Portuguese",
    "ru": "Russian",
    "zh-cn": "Chinese (Simplified)",
    "zh-tw": "Chinese (Traditional)",
}

SUPPORTED_FORMATS: list[str] = ["xlsx", "csv", "json"]

# Mapping from short codes to API language codes
SHORT_TO_API_CODE: dict[str, str] = {
    "de": "de-DE",
    "en": "en-EN",
    "es": "es-ES",
    "fr": "fr-FR",
    "it": "it-IT",
    "ja": "ja-JP",
    "ko": "ko-KR",
    "pl": "pl-PL",
    "pt": "pt-BR",
    "ru": "ru-RU",
    "zh-cn": "zh-Hans",
    "zh-tw": "zh-Hant",
}

# Field names for export (internal keys used in card dictionaries)
EXPORT_FIELD_NAMES: list[str] = [
    "Nation",
    "Name",
    "Type",
    "Rarity",
    "Abilities",
    "Set",
    "Quantity",
    "Credits",
    "Attack",
    "Defense",
    "Description",
]

# Logging setup
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)


class TranslationLoader:
    """Loads translations dynamically from KARDS website JS files."""

    BASE_URL = "https://www.kards.com"

    # Known mappings (fallback if dynamic loading fails)
    KNOWN_MAPPINGS: dict[str, dict[str, str]] = {
        "type": {
            "infantry": "llkqn9",
            "tank": "QIFcAI",
            "armor": "QIFcAI",
            "artillery": "ziY9j1",
            "fighter": "al73ht",
            "air": "al73ht",
            "order": "UYUgdb",
            "countermeasure": "qM208o",
        },
        "faction": {
            "Soviet": "iROGPL",
            "USA": "Mqy/Zy",
            "Japan": "A1ET6E",
            "Germany": "XTtR6a",
            "Britain": "OICU0U",
            "France": "+gY+iO",
            "Italy": "MFljzs",
            "Poland": "sfwBnA",
        },
        "rarity": {
            "Standard": "TJBHlP",
            "Limited": "HhURN3",
            "Special": "qBFI6F",
            "Elite": "JEzmqf",
        },
        "set": {
            "Base": "Nzwli2",
            "Allegiance": "bPobF4",
            "TheatersOfWar": "MPVNE8",
            "Breakthrough": "paHq3y",
            "WorldAtWar": "tkXxPO",
            "CovertOps": "/Adfjf",
            "BloodAndIron": "vhFlLC",
            "Legions": "a6nh/L",
            "NavalWarfare": "6bDKSi",
            "Homefront": "5rE6vr",
            "WinterWar": "wDZgXG",
        },
        "header": {
            "Name": "6YtxFj",
            "Nation": "EpPhw3",
            "Type": "+zy2Nq",
            "Cost": "I99Miw",
            "Rarity": "U86mdr",
            "Attack": "H+HHXa",
            "Defense": "nvGuxv",
            "Set": "NfZUor",
            "Credits": "TgRgc0",
        },
    }

    # Language index in translation arrays (order in JS file)
    LANG_INDEX: dict[str, int] = {
        "de": 0,
        "en": 1,
        "es": 2,
        "fr": 3,
        "it": 4,
        "ja": 5,
        "ko": 6,
        "pl": 7,
        "pt": 8,
        "ru": 9,
        "zh-tw": 10,
        "zh-cn": 11,
    }

    # Header translations (KARDS site uses complex phrases, so we define simple ones)
    HEADER_TRANSLATIONS: dict[str, dict[str, str]] = {
        "Name": {
            "de": "Name", "en": "Name", "es": "Nombre", "fr": "Nom",
            "it": "Nome", "ja": "名前", "ko": "이름", "pl": "Nazwa",
            "pt": "Nome", "ru": "Название", "zh-tw": "名稱", "zh-cn": "名称",
        },
        "Nation": {
            "de": "Nation", "en": "Nation", "es": "Nación", "fr": "Nation",
            "it": "Nazione", "ja": "国家", "ko": "국가", "pl": "Nacja",
            "pt": "Nação", "ru": "Нация", "zh-tw": "國家", "zh-cn": "国家",
        },
        "Type": {
            "de": "Typ", "en": "Type", "es": "Tipo", "fr": "Type",
            "it": "Tipo", "ja": "タイプ", "ko": "유형", "pl": "Typ",
            "pt": "Tipo", "ru": "Тип", "zh-tw": "類型", "zh-cn": "类型",
        },
        "Rarity": {
            "de": "Seltenheit", "en": "Rarity", "es": "Rareza", "fr": "Rareté",
            "it": "Rarità", "ja": "レアリティ", "ko": "희귀도", "pl": "Rzadkość",
            "pt": "Raridade", "ru": "Редкость", "zh-tw": "稀有度", "zh-cn": "稀有度",
        },
        "Attack": {
            "de": "Angriff", "en": "Attack", "es": "Ataque", "fr": "Attaque",
            "it": "Attacco", "ja": "攻撃", "ko": "공격", "pl": "Atak",
            "pt": "Ataque", "ru": "Атака", "zh-tw": "攻擊", "zh-cn": "攻击",
        },
        "Defense": {
            "de": "Verteidigung", "en": "Defense", "es": "Defensa", "fr": "Défense",
            "it": "Difesa", "ja": "防御", "ko": "방어", "pl": "Obrona",
            "pt": "Defesa", "ru": "Защита", "zh-tw": "防禦", "zh-cn": "防御",
        },
        "Abilities": {
            "de": "Fähigkeiten", "en": "Abilities", "es": "Habilidades",
            "fr": "Capacités", "it": "Abilità", "ja": "能力", "ko": "능력",
            "pl": "Umiejętności", "pt": "Habilidades", "ru": "Способности",
            "zh-tw": "能力", "zh-cn": "能力",
        },
        "Set": {
            "de": "Set", "en": "Set", "es": "Set", "fr": "Set",
            "it": "Set", "ja": "セット", "ko": "세트", "pl": "Zestaw",
            "pt": "Coleção", "ru": "Сет", "zh-tw": "系列", "zh-cn": "系列",
        },
        "Quantity": {
            "de": "Anzahl", "en": "Quantity", "es": "Cantidad", "fr": "Quantité",
            "it": "Quantità", "ja": "数量", "ko": "수량", "pl": "Ilość",
            "pt": "Quantidade", "ru": "Количество", "zh-tw": "數量", "zh-cn": "数量",
        },
        "Credits": {
            "de": "Kredits", "en": "Credits", "es": "Créditos", "fr": "Crédits",
            "it": "Crediti", "ja": "クレジット", "ko": "크레딧", "pl": "Kredyty",
            "pt": "Créditos", "ru": "Кредиты", "zh-tw": "點數", "zh-cn": "点数",
        },
        "Description": {
            "de": "Beschreibung", "en": "Description", "es": "Descripción",
            "fr": "Description", "it": "Descrizione", "ja": "説明", "ko": "설명",
            "pl": "Opis", "pt": "Descrição", "ru": "Описание",
            "zh-tw": "描述", "zh-cn": "描述",
        },
    }

    def __init__(self) -> None:
        """Initialize translation loader."""
        self.translations: dict[str, str] = {}
        self.mappings: dict[str, dict[str, str]] = self.KNOWN_MAPPINGS.copy()
        self.language: str = "en"

    async def load_translations(self, language: str) -> None:
        """Load translations for specified language from KARDS JS files.

        Args:
            language: Target language code (e.g., 'ru', 'en')
        """
        self.language = language
        logger.info(f"Loading translations for {language}...")

        try:
            async with httpx.AsyncClient(timeout=30.0) as client:
                # First, get the main page to find JS file URLs
                response = await client.get(f"{self.BASE_URL}/{language}/decks/collection")
                html = response.text

                # Find translation JS file URL
                js_urls = re.findall(r'/_next/static/chunks/[^"\']+\.js', html)

                translation_content = ""
                for js_url in js_urls:
                    if "2840-" in js_url:
                        js_response = await client.get(f"{self.BASE_URL}{js_url}")
                        translation_content = js_response.text
                        break

                if translation_content:
                    self._parse_translations(translation_content, language)
                    logger.info(f"Loaded {len(self.translations)} translation keys")
                else:
                    logger.warning("Translation JS file not found, using fallback")

        except Exception as e:
            logger.warning(f"Failed to load translations dynamically: {e}")
            logger.info("Using fallback translations")

    def _parse_translations(self, js_content: str, language: str) -> None:
        """Parse translations from JS content.

        Args:
            js_content: JavaScript file content
            language: Target language code
        """
        # Extract all translation key-value pairs for the target language
        # Format in JS: "key":"value" repeated for each language
        lang_idx = self.LANG_INDEX.get(language, 9)  # Default to Russian

        # Find all translation IDs we need
        all_ids: set[str] = set()
        for category_mappings in self.mappings.values():
            all_ids.update(category_mappings.values())

        # Extract translations for each ID
        for trans_id in all_ids:
            # Find all occurrences of this ID with its value
            pattern = re.compile(rf'"{re.escape(trans_id)}":"([^"]*)"')
            matches = pattern.findall(js_content)

            if matches and lang_idx < len(matches):
                self.translations[trans_id] = matches[lang_idx]
            elif matches:
                # Use last match as fallback
                self.translations[trans_id] = matches[-1]

    def translate(self, category: str, value: str) -> str:
        """Translate a value using loaded translations.

        Args:
            category: Category name (type, faction, rarity, set)
            value: Original value from API

        Returns:
            Translated value or original if not found
        """
        if not value:
            return ""

        # Normalize value for lookup
        normalized = value.strip()

        # Get mapping ID for this value
        category_mappings = self.mappings.get(category, {})
        trans_id = category_mappings.get(normalized)

        if not trans_id:
            # Try case-insensitive lookup
            for key, tid in category_mappings.items():
                if key.lower() == normalized.lower():
                    trans_id = tid
                    break

        if trans_id and trans_id in self.translations:
            translated = self.translations[trans_id]
            # Clean up translation (remove surrounding quotes if present)
            translated = self._strip_quotes(translated)
            return translated

        # Return original value if no translation found
        return normalized

    @staticmethod
    def _strip_quotes(text: str) -> str:
        """Remove surrounding quotes from text.

        Args:
            text: Text possibly wrapped in quotes

        Returns:
            Text with surrounding quotes removed
        """
        if not text:
            return text
        # Remove unicode quotes «» and regular quotes
        quote_pairs = [('«', '»'), ('"', '"'), ("'", "'"), ('"', '"'), (''', ''')]
        for open_q, close_q in quote_pairs:
            if text.startswith(open_q) and text.endswith(close_q):
                text = text[len(open_q):-len(close_q)]
        return text

    @staticmethod
    def sanitize_text(text: str) -> str:
        """Sanitize text by removing quotes, newlines, and decoding escapes.

        Args:
            text: Text to sanitize

        Returns:
            Sanitized text with quotes removed, newlines replaced with spaces,
            escape sequences decoded, and duplicate spaces removed
        """
        if not text:
            return text
        # Decode escape sequences (e.g., \xab → «, \xdf → ß)
        try:
            text = text.encode('latin-1').decode('unicode_escape').encode(
                'latin-1'
            ).decode('utf-8')
        except (UnicodeDecodeError, UnicodeEncodeError):
            pass  # Keep original if decoding fails
        # Remove surrounding quotes
        text = TranslationLoader._strip_quotes(text)
        # Replace newlines with space
        text = text.replace('\r\n', ' ').replace('\n', ' ')
        # Remove duplicate spaces
        text = ' '.join(text.split())
        return text

    def translate_headers(self, headers: list[str]) -> list[str]:
        """Translate column headers using predefined translations.

        Args:
            headers: List of English header names

        Returns:
            List of translated headers (falls back to English if not found)
        """
        result = []
        for header in headers:
            if header in self.HEADER_TRANSLATIONS:
                lang_translations = self.HEADER_TRANSLATIONS[header]
                result.append(lang_translations.get(self.language, header))
            else:
                result.append(header)
        return result


def extract_localized_field(
    field_data: dict[str, str] | str,
    language: str,
    field_name: str = "",
) -> str:
    """Extract localized field value with fallback logic.

    Priority: selected language -> English -> first available

    Args:
        field_data: Dictionary with language codes as keys or string value
        language: Target language code (short code like 'en', 'ru', 'ja')
        field_name: Field name for debug logging

    Returns:
        Localized string value or empty string if not found
    """
    if not isinstance(field_data, dict):
        return str(field_data) if field_data else ""

    # Try short code first (e.g., 'ru')
    if language in field_data:
        return field_data[language]

    # Try API code (e.g., 'ru-RU')
    api_code = SHORT_TO_API_CODE.get(language, language)
    if api_code in field_data:
        return field_data[api_code]

    # Fallback to English (try both short and API codes)
    if "en" in field_data:
        logger.debug(
            f"{field_name} not available in {language}, using English fallback"
        )
        return field_data["en"]

    if "en-EN" in field_data:
        logger.debug(
            f"{field_name} not available in {language}, using English fallback"
        )
        return field_data["en-EN"]

    # Fallback to first available language
    for lang_code in SUPPORTED_LANGUAGES:
        if lang_code in field_data:
            logger.debug(
                f"{field_name} not available in {language}, using {lang_code}"
            )
            return field_data[lang_code]
        # Also check API code
        lang_api_code = SHORT_TO_API_CODE.get(lang_code, lang_code)
        if lang_api_code in field_data:
            logger.debug(
                f"{field_name} not available in {language}, using {lang_api_code}"
            )
            return field_data[lang_api_code]

    return ""


def export_to_xlsx(
    cards: list[dict[str, str]],
    filename: str,
    language: str,
    translator: TranslationLoader | None = None,
) -> None:
    """Export cards to Excel format with formatting.

    Args:
        cards: List of card dictionaries
        filename: Output filename
        language: Language code for metadata
        translator: Optional translator for localizing headers
    """
    logger.info(f"Exporting {len(cards)} cards to Excel: {filename}")

    wb = Workbook()
    ws = wb.active
    if not ws:
        msg = "Failed to create worksheet"
        raise RuntimeError(msg)
    ws.title = "KARDS Cards"

    # Translate headers if translator is available
    if translator:
        headers = translator.translate_headers(EXPORT_FIELD_NAMES)
    else:
        headers = list(EXPORT_FIELD_NAMES)

    ws.append(headers)

    # Header formatting
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Data rows
    sorted_cards = sorted(cards, key=lambda x: x.get("Name", ""))
    for card in sorted_cards:
        row = [card.get(field, "") for field in EXPORT_FIELD_NAMES]
        ws.append(row)

    # Column widths
    # Nation, Name, Type, Rarity, Abilities, Set, Quantity, Credits, Attack, Defense, Description
    column_widths = [15, 35, 18, 15, 12, 20, 10, 10, 8, 8, 60]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(cards) + 1}"

    wb.save(filename)
    logger.info(f"Excel file created: {filename} ({len(cards)} cards)")


def export_to_csv(
    cards: list[dict[str, str]],
    filename: str,
    translator: TranslationLoader | None = None,
) -> None:
    """Export cards to CSV format.

    Uses UTF-8 with BOM for Excel compatibility on Windows.

    Args:
        cards: List of card dictionaries
        filename: Output filename
        translator: Optional translator for localizing headers
    """
    logger.info(f"Exporting {len(cards)} cards to CSV: {filename}")

    # Translate headers if translator is available
    if translator:
        headers = translator.translate_headers(EXPORT_FIELD_NAMES)
    else:
        headers = list(EXPORT_FIELD_NAMES)

    with open(filename, "w", newline="", encoding="utf-8-sig") as csvfile:
        writer = csv.writer(csvfile)
        writer.writerow(headers)

        sorted_cards = sorted(cards, key=lambda x: x.get("Name", ""))
        for card in sorted_cards:
            row = [card.get(field, "") for field in EXPORT_FIELD_NAMES]
            writer.writerow(row)

    logger.info(f"CSV file created: {filename} ({len(cards)} cards)")


def export_to_json(cards: list[dict[str, str]], filename: str, language: str) -> None:
    """Export cards to JSON format with metadata.

    Args:
        cards: List of card dictionaries
        filename: Output filename
        language: Language code for metadata
    """
    logger.info(f"Exporting {len(cards)} cards to JSON: {filename}")

    sorted_cards = sorted(cards, key=lambda x: x.get("Name", ""))

    output_data = {
        "metadata": {
            "language": language,
            "language_name": SUPPORTED_LANGUAGES[language],
            "total_cards": len(cards),
        },
        "cards": sorted_cards,
    }

    with open(filename, "w", encoding="utf-8") as jsonfile:
        json.dump(output_data, jsonfile, ensure_ascii=False, indent=2)

    logger.info(f"JSON file created: {filename} ({len(cards)} cards)")


class KardsFinalScraper:
    def __init__(
        self,
        language: str = "ru",
        export_format: str = "csv",
        output_file: str | None = None,
    ) -> None:
        """Initialize KARDS scraper.

        Args:
            language: Language code for card text (default: ru)
            export_format: Export format - xlsx, csv, or json (default: csv)
            output_file: Custom output filename (optional, auto-generated if None)

        Raises:
            ValueError: If language or format is not supported
        """
        if language not in SUPPORTED_LANGUAGES:
            msg = (
                f"Unsupported language: {language}. "
                f"Supported: {', '.join(SUPPORTED_LANGUAGES.keys())}"
            )
            raise ValueError(msg)

        if export_format not in SUPPORTED_FORMATS:
            msg = (
                f"Unsupported format: {export_format}. "
                f"Supported: {', '.join(SUPPORTED_FORMATS)}"
            )
            raise ValueError(msg)

        self.language = language
        self.export_format = export_format
        self.output_file = output_file

        self.url = f"https://www.kards.com/{language}/decks/collection"
        self.cards: list[dict[str, str]] = []
        self.card_ids: set[str] = set()
        self.browser: Browser | None = None
        self.page: Page | None = None
        self.api_data: list[dict[str, Any]] = []
        self.translator: TranslationLoader = TranslationLoader()

    async def initialize_browser(self) -> None:
        """Initialize Playwright browser."""
        logger.info("Initializing browser...")
        self.playwright = await async_playwright().start()
        self.browser = await self.playwright.chromium.launch(headless=True)
        self.page = await self.browser.new_page()

        # Listen to server responses
        async def on_response(response: Any) -> None:
            url = response.url
            # Look for API requests
            if ('graphql' in url.lower() or 'api' in url.lower()) and response.ok:
                try:
                    text = await response.text()
                    if text:
                        try:
                            data = json.loads(text)
                            if 'data' in data or 'errors' in data:
                                self.api_data.append(data)
                                logger.info(f"API data received: {len(text)} bytes")
                        except json.JSONDecodeError:
                            pass
                except Exception as e:
                    logger.debug(f"Error reading response: {e}")

        self.page.on("response", on_response)

        logger.info("Browser initialized successfully")

    async def close_browser(self) -> None:
        """Close browser."""
        if self.page:
            await self.page.close()
        if self.browser:
            await self.browser.close()
        if self.playwright:
            await self.playwright.stop()
        logger.info("Browser closed")

    async def load_page(self) -> None:
        """Load page."""
        logger.info(f"Loading page {self.url}...")
        try:
            if not self.page:
                msg = "Page not initialized"
                raise RuntimeError(msg)
            await self.page.goto(self.url, wait_until="networkidle", timeout=60000)
            logger.info("Page loaded")
            await self.page.wait_for_timeout(2000)  # Additional wait for data loading
        except Exception as e:
            logger.error(f"Error loading page: {e}")
            raise

    async def load_all_cards(self) -> None:
        """Load all cards by clicking 'LOAD MORE' button."""
        logger.info("Loading all cards by clicking 'LOAD MORE'...")

        if not self.page:
            msg = "Page not initialized"
            raise RuntimeError(msg)

        load_more_clicks = 0
        max_clicks = 50

        while load_more_clicks < max_clicks:
            # Find "LOAD MORE" button
            buttons = await self.page.query_selector_all('button')

            load_more_button = None
            for button in buttons:
                try:
                    text = await button.inner_text()
                    if "LOAD MORE" in text.upper():
                        load_more_button = button
                        break
                except Exception:
                    pass

            if not load_more_button:
                logger.info("'LOAD MORE' button not found. All cards loaded.")
                break

            # Scroll to button and click
            try:
                await load_more_button.scroll_into_view_if_needed()
                await asyncio.sleep(0.3)
                await load_more_button.click()
                load_more_clicks += 1
                logger.info(f"Click on 'LOAD MORE' {load_more_clicks}/{max_clicks}")

                # Wait for new cards to load
                await asyncio.sleep(2)

            except Exception as e:
                logger.warning(f"Error clicking LOAD MORE: {e}")
                break

        logger.info(f"Loading completed after {load_more_clicks} clicks")

    def parse_api_data(self) -> None:
        """Parse API data to extract cards."""
        logger.info(f"Parsing {len(self.api_data)} API responses...")

        for api_response in self.api_data:
            try:
                # Look for card data in response
                if 'data' in api_response and isinstance(api_response['data'], dict):
                    data = api_response['data']

                    # GraphQL can return data in different structures
                    for key, value in data.items():
                        if isinstance(value, dict) and 'edges' in value:
                            # This looks like a list of cards
                            edges = value.get('edges', [])
                            for edge in edges:
                                if 'node' in edge:
                                    self._process_card(edge['node'])

            except Exception as e:
                logger.warning(f"Error parsing response: {e}")

        logger.info(f"Total cards extracted: {len(self.cards)}")

    def _process_card(self, card_node: dict[str, Any]) -> None:
        """Process card node from API."""
        try:
            card_id = card_node.get('cardId', '')
            if card_id in self.card_ids:
                return  # Already processed

            # Get card JSON data
            json_data = card_node.get('json', {})

            # Get title in target language
            title = extract_localized_field(
                json_data.get('title', {}), self.language, 'title'
            )

            if not title:
                return

            # Get raw values from API
            faction_raw = str(json_data.get('faction', ''))
            type_raw = str(json_data.get('type', ''))
            rarity_raw = str(json_data.get('rarity', ''))
            set_raw = str(json_data.get('set', ''))

            # Get description in target language
            description = extract_localized_field(
                json_data.get('text', {}), self.language, 'description'
            )

            # Translate and sanitize values using loaded translations
            sanitize = TranslationLoader.sanitize_text
            card_info: dict[str, str] = {
                'Nation': sanitize(self.translator.translate('faction', faction_raw)),
                'Name': sanitize(title),
                'Type': sanitize(self.translator.translate('type', type_raw)),
                'Rarity': sanitize(self.translator.translate('rarity', rarity_raw)),
                'Abilities': '',
                'Set': sanitize(self.translator.translate('set', set_raw)),
                'Quantity': '',
                'Credits': str(json_data.get('kredits', '')),
                'Attack': str(json_data.get('attack', '')),
                'Defense': str(json_data.get('defense', '')),
                'Description': sanitize(description),
            }

            self.cards.append(card_info)
            self.card_ids.add(card_id)

        except Exception as e:
            logger.warning(f"Error processing card: {e}")

    def _generate_filename(self) -> str:
        """Generate output filename based on language and format.

        Returns:
            Filename like 'kards_collection_en.csv'
        """
        return f"kards_collection_{self.language}.{self.export_format}"

    async def run(self) -> bool:
        """Main scraper entry point."""
        try:
            logger.info("╔════════════════════════════════════════════════════════════╗")
            logger.info("║       KARDS Card Scraper - Multi-language Export            ║")
            logger.info("╚════════════════════════════════════════════════════════════╝")
            logger.info(f"Language: {SUPPORTED_LANGUAGES[self.language]}")
            logger.info(f"Export format: {self.export_format.upper()}")
            logger.info("")

            # Load translations for localized field values
            await self.translator.load_translations(self.language)

            await self.initialize_browser()
            await self.load_page()
            await self.load_all_cards()

            logger.info(f"\nReceived {len(self.api_data)} API responses")
            self.parse_api_data()

            if not self.cards:
                logger.warning("No cards found")
                return False

            filename = self.output_file or self._generate_filename()

            # Export using appropriate function
            if self.export_format == "xlsx":
                export_to_xlsx(self.cards, filename, self.language, self.translator)
            elif self.export_format == "csv":
                export_to_csv(self.cards, filename, self.translator)
            elif self.export_format == "json":
                export_to_json(self.cards, filename, self.language)

            logger.info("\nScraper completed successfully!")
            return True

        except Exception as e:
            logger.error(f"Critical error: {e}")
            logger.error(traceback.format_exc())
            return False

        finally:
            await self.close_browser()


__version__ = "0.1"


async def main() -> None:
    """Entry point for CLI."""
    parser = argparse.ArgumentParser(
        description="KARDS Card Scraper - Export cards to various formats",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=f"""
Supported languages:
{chr(10).join(f'  {code:10} - {name}' for code, name in SUPPORTED_LANGUAGES.items())}

Supported formats:
  csv        - Comma-separated values (default)
  xlsx       - Excel spreadsheet
  json       - JSON with metadata

Examples:
  %(prog)s -l en               # English cards in CSV format
  %(prog)s -l ja -f json       # Japanese cards in JSON format
  %(prog)s -l ru -f xlsx -o my_cards.xlsx
        """,
    )

    parser.add_argument(
        "--version",
        "-v",
        action="version",
        version=f"%(prog)s {__version__}",
    )

    parser.add_argument(
        "--language",
        "-l",
        type=str,
        required=True,
        choices=list(SUPPORTED_LANGUAGES.keys()),
        help="Language for card text (required)",
    )

    parser.add_argument(
        "--format",
        "-f",
        type=str,
        default="csv",
        choices=SUPPORTED_FORMATS,
        help="Export format (default: csv)",
    )

    parser.add_argument(
        "--output",
        "-o",
        type=str,
        default=None,
        help="Output filename (default: auto-generated)",
    )

    # Show help if no arguments provided
    if len(sys.argv) == 1:
        parser.print_help()
        sys.exit(0)

    args = parser.parse_args()

    scraper = KardsFinalScraper(
        language=args.language,
        export_format=args.format,
        output_file=args.output,
    )

    success = await scraper.run()
    sys.exit(0 if success else 1)


if __name__ == "__main__":
    asyncio.run(main())
