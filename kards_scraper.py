#!/usr/bin/env python3
"""
KARDS Card Scraper
Извлекает список карт с https://www.kards.com/ru/decks/collection
и экспортирует в Excel формат
"""

import asyncio
import logging
from typing import List, Dict, Optional
from playwright.async_api import async_playwright, Page, Browser
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from datetime import datetime
import sys

# Настройка логирования
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)


class KardsScraper:
    def __init__(self):
        self.url = "https://www.kards.com/ru/decks/collection"
        self.cards: List[Dict] = []
        self.browser: Optional[Browser] = None
        self.page: Optional[Page] = None

    async def initialize_browser(self):
        """Инициализация Playwright браузера"""
        logger.info("Инициализация браузера...")
        self.playwright = await async_playwright().start()
        self.browser = await self.playwright.chromium.launch(headless=True)
        self.page = await self.browser.new_page()
        logger.info("Браузер успешно инициализирован")

    async def close_browser(self):
        """Закрытие браузера"""
        if self.page:
            await self.page.close()
        if self.browser:
            await self.browser.close()
        if self.playwright:
            await self.playwright.stop()
        logger.info("Браузер закрыт")

    async def load_page(self):
        """Загрузка страницы и ожидание полной загрузки карт"""
        logger.info(f"Загрузка страницы {self.url}...")
        try:
            await self.page.goto(self.url, wait_until="networkidle", timeout=60000)
            logger.info("Страница загружена")
        except Exception as e:
            logger.error(f"Ошибка при загрузке страницы: {e}")
            raise

    async def wait_for_cards(self):
        """Ожидание появления карт на странице"""
        logger.info("Ожидание загрузки карт...")
        try:
            # Ждем появления контейнера с картами
            await self.page.wait_for_selector('[class*="card"]', timeout=30000)
            logger.info("Карты найдены на странице")
        except Exception as e:
            logger.warning(f"Не удалось дождаться загрузки карт: {e}")

    async def scroll_and_load_all_cards(self):
        """Загрузка всех карт путем нажатия на кнопку 'LOAD MORE'"""
        logger.info("Загрузка всех карт путем нажатия 'LOAD MORE'...")

        load_more_clicks = 0
        max_clicks = 100  # Максимум кликов на LOAD MORE

        while load_more_clicks < max_clicks:
            # Ищем кнопку "LOAD MORE"
            load_more_button = await self.page.query_selector('button:has-text("LOAD MORE"), button:has-text("Load More"), a:has-text("LOAD MORE")')

            if not load_more_button:
                logger.info("Кнопка 'LOAD MORE' не найдена. Все карты загружены.")
                break

            # Прокручиваем к кнопке
            await load_more_button.scroll_into_view_if_needed()
            await asyncio.sleep(0.5)

            # Нажимаем кнопку
            try:
                await load_more_button.click()
                load_more_clicks += 1
                logger.info(f"Клик на 'LOAD MORE' {load_more_clicks}/{max_clicks}")

                # Ждем загрузки новых карт
                await asyncio.sleep(2)

                # Проверяем, что карты загружены
                cards_count = await self.page.evaluate("document.querySelectorAll('[class*=\"Card_card\"]').length")
                logger.info(f"Загружено карт: {cards_count}")

            except Exception as e:
                logger.warning(f"Ошибка при клике на LOAD MORE: {e}")
                break

        logger.info("Загрузка карт завершена")

    async def extract_cards(self):
        """Извлечение данных всех карт со страницы"""
        logger.info("Извлечение данных карт...")

        try:
            # Получаем структурированные данные карт с помощью JavaScript
            cards_data = await self.page.evaluate("""
                () => {
                    const cards = [];

                    // Используем селектор, который мы найдли при инспекции
                    const cardElements = document.querySelectorAll('[class*="Card_card__"]');

                    cardElements.forEach(element => {
                        try {
                            const cardInfo = {};

                            // Получаем весь текст элемента
                            const fullText = element.innerText || element.textContent;

                            // Разбираем текст на строки
                            if (fullText && fullText.trim().length > 0) {
                                const lines = fullText.trim().split('\\n').filter(l => l.trim());

                                // Первая строка - название карты
                                if (lines.length > 0) {
                                    cardInfo.title = lines[0].trim();
                                }

                                // Ищем информацию о стоимости (обычно начинается с числа и содержит "kredits" или символ)
                                cardInfo.cost = lines.find(l => /^[0-9]+/.test(l.trim())) || '';

                                // Ищем информацию о редкости
                                const rarityKeywords = ['Common', 'Limited', 'Special', 'Elite', 'Rare', 'Обычная', 'Ограниченная', 'Особая', 'Элита', 'Редкая'];
                                cardInfo.rarity = lines.find(l => rarityKeywords.some(k => l.includes(k))) || '';

                                // Ищем информацию о типе (Unit, Order, Countermeasure)
                                const typeKeywords = ['Unit', 'Order', 'Countermeasure', 'Юнит', 'Приказ', 'Контрмера'];
                                cardInfo.type = lines.find(l => typeKeywords.some(k => l.includes(k))) || '';

                                cardInfo.fullText = fullText.trim();
                                cardInfo.html = element.className || '';
                                cards.push(cardInfo);
                            }
                        } catch (e) {
                            console.error('Ошибка при обработке элемента:', e);
                        }
                    });

                    return cards;
                }
            """)

            logger.info(f"Найдено потенциальных элементов: {len(cards_data)}")

            # Обработка извлеченных данных
            seen_titles = set()
            for card_data in cards_data:
                title = card_data.get('title', '').strip()
                if title and title not in seen_titles:
                    card_info = {
                        'Название': title,
                        'Страна': '',
                        'Тип': card_data.get('type', '').strip(),
                        'Стоимость': card_data.get('cost', '').strip(),
                        'Редкость': card_data.get('rarity', '').strip(),
                        'Атака': '',
                        'Защита': '',
                        'Набор': '',
                        'Описание': card_data.get('fullText', '').strip()
                    }

                    self.cards.append(card_info)
                    seen_titles.add(title)

            logger.info(f"Успешно извлечено уникальных карт: {len(self.cards)}")

        except Exception as e:
            logger.error(f"Ошибка при извлечении данных карт: {e}")
            import traceback
            logger.error(traceback.format_exc())

    def _parse_card_data(self, text: str) -> Optional[Dict]:
        """Парсинг текста карты в структурированные данные"""
        lines = [line.strip() for line in text.split('\n') if line.strip()]

        if not lines:
            return None

        card_info = {
            'Название': '',
            'Страна': '',
            'Тип': '',
            'Стоимость': '',
            'Редкость': '',
            'Атака': '',
            'Защита': '',
            'Набор': '',
            'Описание': ''
        }

        # Определяем редкость по ключевым словам
        rarities = ['Common', 'Limited', 'Special', 'Elite', 'Rare', 'Обычная', 'Ограниченная', 'Особая', 'Элита', 'Редкая']
        countries = ['Германия', 'Британия', 'СССР', 'Япония', 'США', 'Germany', 'Britain', 'Soviet Union', 'Japan', 'USA']
        card_types = ['Unit', 'Order', 'Countermeasure', 'Юнит', 'Приказ', 'Контрмера']

        card_info['Название'] = lines[0] if lines else ''

        # Простой парсинг доступной информации
        for line in lines[1:]:
            for country in countries:
                if country in line:
                    card_info['Страна'] = country
                    break

            for card_type in card_types:
                if card_type in line:
                    card_info['Тип'] = card_type
                    break

            for rarity in rarities:
                if rarity in line:
                    card_info['Редкость'] = rarity
                    break

            # Поиск стоимости (kredits, очки)
            if 'kredits' in line.lower() or 'очк' in line.lower():
                card_info['Стоимость'] = line

            # Поиск атаки/защиты
            if 'атак' in line.lower() or 'attack' in line.lower():
                card_info['Атака'] = line
            elif 'защит' in line.lower() or 'defense' in line.lower():
                card_info['Защита'] = line

        # Описание - все остальное
        if len(lines) > 1:
            card_info['Описание'] = ' '.join(lines[1:])

        return card_info

    def export_to_excel(self, filename: str = "kards_cards.xlsx"):
        """Экспорт данных в Excel формат"""
        logger.info(f"Экспорт данных в Excel ({filename})...")

        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "KARDS Cards"

            # Заголовки колонок
            headers = ['Название', 'Страна', 'Тип', 'Стоимость', 'Редкость', 'Атака', 'Защита', 'Набор', 'Описание']
            ws.append(headers)

            # Форматирование заголовков
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")

            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

            # Добавление данных карт
            for card in self.cards:
                row = [
                    card.get('Название', ''),
                    card.get('Страна', ''),
                    card.get('Тип', ''),
                    card.get('Стоимость', ''),
                    card.get('Редкость', ''),
                    card.get('Атака', ''),
                    card.get('Защита', ''),
                    card.get('Набор', ''),
                    card.get('Описание', '')
                ]
                ws.append(row)

            # Автоширина колонок
            column_widths = [30, 15, 15, 15, 15, 10, 10, 20, 50]
            for i, width in enumerate(column_widths, 1):
                ws.column_dimensions[get_column_letter(i)].width = width

            # Заморозка первой строки
            ws.freeze_panes = "A2"

            # Добавление фильтров
            ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(self.cards) + 1}"

            # Сохранение файла
            wb.save(filename)
            logger.info(f"Excel файл успешно создан: {filename}")
            logger.info(f"Всего карт в файле: {len(self.cards)}")

        except Exception as e:
            logger.error(f"Ошибка при экспорте в Excel: {e}")
            raise

    async def run(self):
        """Основной метод запуска скрейпера"""
        try:
            logger.info("Запуск KARDS Card Scraper...")
            logger.info(f"Целевая страница: {self.url}")

            await self.initialize_browser()
            await self.load_page()
            await self.wait_for_cards()
            await self.scroll_and_load_all_cards()
            await self.extract_cards()

            self.export_to_excel()

            logger.info("Скрейпер успешно завершил работу!")
            return True

        except Exception as e:
            logger.error(f"Критическая ошибка: {e}")
            return False

        finally:
            await self.close_browser()


async def main():
    """Точка входа программы"""
    scraper = KardsScraper()
    success = await scraper.run()
    sys.exit(0 if success else 1)


if __name__ == "__main__":
    asyncio.run(main())
