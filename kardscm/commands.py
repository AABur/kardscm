"""Business logic for CLI commands."""

from __future__ import annotations

import logging
import sqlite3
from datetime import UTC, datetime
from pathlib import Path

from openpyxl import load_workbook

from kardscm.advisor import build_analysis_context, get_llm_response
from kardscm.config import get_advisor_config, get_language_config
from kardscm.constants import DEFAULT_DB_PATH
from kardscm.export import (
    add_deck_sheet,
    export_deck_to_json,
    export_to_csv,
    export_to_json,
    export_to_xlsx,
    translate_card_for_export,
)
from kardscm.helpers import parse_int
from kardscm.importing import parse_deck_file
from kardscm.models import DeckCardEntry
from kardscm.scraping import scrape_cards
from kardscm.storage import (
    compute_deck_stats,
    delete_all_decks,
    delete_deck,
    fetch_all_decks,
    fetch_cards,
    fetch_deck_cards,
    find_card_id,
    find_card_id_by_exile,
    find_deck_by_name,
    get_card_quantity_by_id,
    get_connection,
    initialize_schema,
    insert_deck,
    insert_deck_cards,
    insert_match,
    set_metadata,
    update_card_quantity_by_id,
    update_quantity,
    upsert_cards,
)

logger = logging.getLogger(__name__)


def _utc_timestamp() -> str:
    return datetime.now(UTC).replace(tzinfo=None).isoformat(timespec="seconds")


def validate_file(path: str, expected_ext: str, must_exist: bool = False) -> Path:
    """Validate file extension and existence.

    Args:
        path: File path string.
        expected_ext: Expected extension (e.g. ".xlsx").
        must_exist: If True, file must exist on disk.

    Returns:
        Resolved Path object.

    Raises:
        SystemExit: If validation fails.
    """
    p = Path(path)
    if p.suffix != expected_ext:
        raise SystemExit(f"Expected {expected_ext} file, got: {p.suffix or '(no extension)'}")
    if must_exist and not p.exists():
        raise SystemExit(f"File not found: {path}")
    return p


def _read_xlsx_quantities(filename: str) -> list[tuple[str, str, int | None]]:
    """Read faction, title, quantity from XLSX file.

    Column names are determined by the configured language.
    Returns (faction_display, localized_title, quantity) tuples.

    Args:
        filename: Path to XLSX file.

    Returns:
        List of (faction, title, quantity) tuples.

    Raises:
        FileNotFoundError: If file doesn't exist.
        ValueError: If required columns missing.
    """
    if not Path(filename).exists():
        raise FileNotFoundError(f"File not found: {filename}")

    lang_config = get_language_config()
    headers_list = lang_config.export_headers
    header_map = {
        headers_list[0]: "faction",
        headers_list[1]: "title",
        headers_list[6]: "quantity",
    }

    wb = load_workbook(filename)
    ws = wb.active
    if not ws:
        raise ValueError("No active worksheet found")

    headers = {
        header_map[cell.value]: col_idx
        for col_idx, cell in enumerate(ws[1], 1)
        if cell.value in header_map
    }
    if len(headers) != len(header_map):
        expected = ", ".join(header_map.keys())
        raise ValueError(f"Missing required columns: {expected}")

    results = []
    for row in ws.iter_rows(min_row=2, values_only=False):
        faction_cell = row[headers["faction"] - 1]
        title_cell = row[headers["title"] - 1]
        qty_cell = row[headers["quantity"] - 1]

        faction = faction_cell.value
        title = title_cell.value
        qty_val = qty_cell.value

        if not faction or not title:
            continue

        qty = parse_int(qty_val)
        results.append((str(faction).strip(), str(title).strip(), qty))

    return results


def sync_collection(db_path: str = DEFAULT_DB_PATH) -> None:
    """Synchronize local SQLite storage with the website.

    Args:
        db_path: SQLite database path.
    """
    lang_config = get_language_config()
    logger.info("Starting sync from website (language: %s)...", lang_config.name)
    cards = scrape_cards()

    with get_connection(db_path) as conn:
        initialize_schema(conn)
        upsert_cards(conn, cards)
        set_metadata(conn, "last_sync", _utc_timestamp())
        set_metadata(conn, "language", lang_config.code)

    logger.info("Sync completed. Stored %s cards.", len(cards))


def export_collection(
    export_format: str,
    filename: str,
    db_path: str = DEFAULT_DB_PATH,
) -> None:
    """Export cards from SQLite to the selected format.

    Args:
        export_format: Export format (csv, json, xlsx).
        filename: Output file path.
        db_path: SQLite database path.
    """
    lang_config = get_language_config()

    with get_connection(db_path) as conn:
        initialize_schema(conn)
        raw_cards = fetch_cards(conn)

    if not raw_cards:
        raise SystemExit("No cards in database. Run 'kards sync' first.")

    cards = [translate_card_for_export(card, lang_config) for card in raw_cards]

    if export_format == "xlsx":
        export_to_xlsx(
            cards,
            filename,
            lang_config.export_headers,
            lang_config.collection_sheet_name,
        )
    elif export_format == "csv":
        export_to_csv(cards, filename, lang_config.export_headers)
    elif export_format == "json":
        export_to_json(cards, filename, lang_config.code, lang_config.name)
    else:
        msg = f"Unsupported format: {export_format}"
        raise ValueError(msg)

    logger.info("Export completed: %s", filename)


def update_collection(filename: str, db_path: str = DEFAULT_DB_PATH) -> None:
    """Update card quantities from XLSX file.

    Args:
        filename: XLSX file path.
        db_path: SQLite database path.
    """
    logger.info("Starting update from file: %s", filename)
    lang_config = get_language_config()

    try:
        updates = _read_xlsx_quantities(filename)
    except (FileNotFoundError, ValueError) as e:
        raise SystemExit(f"Failed to read file: {e}") from e

    if not updates:
        logger.warning("No valid entries found in file")
        return

    # Reverse-map localized faction names to API faction names
    reverse_faction = {v: k for k, v in lang_config.faction_names.items()}

    mapped_updates = []
    for faction_display, title, qty in updates:
        faction_api = reverse_faction.get(faction_display, faction_display)
        mapped_updates.append((faction_api, title, qty))

    with get_connection(db_path) as conn:
        initialize_schema(conn)
        updated_count, not_found = update_quantity(conn, mapped_updates, lang_config.locale_key)

    logger.info(
        "Update completed: %d cards updated, %d not found",
        updated_count,
        len(not_found),
    )

    if not_found:
        for key in not_found:
            logger.warning("Card not found: %s", key)


def import_deck(filename: str, db_path: str = DEFAULT_DB_PATH) -> None:
    """Import a deck from TXT file into the database.

    Args:
        filename: Path to deck TXT file.
        db_path: SQLite database path.
    """
    lang_config = get_language_config()
    logger.info("Importing deck from file: %s", filename)

    try:
        deck = parse_deck_file(filename)
    except (FileNotFoundError, ValueError) as e:
        raise SystemExit(f"Failed to parse deck file: {e}") from e

    with get_connection(db_path) as conn:
        initialize_schema(conn)

        existing = find_deck_by_name(conn, deck["name"])
        if existing:
            raise SystemExit(f"Deck '{deck['name']}' already exists (id={existing['deck_id']})")

        not_found = []
        for card in deck["cards"]:
            faction = lang_config.deck_nation_to_db.get(card["nation"], card["nation"])
            card_id = find_card_id(conn, faction, card["name"], lang_config.locale_key)
            if card_id is None:
                not_found.append(f"{faction} / {card['name']}")

        if not_found:
            lines = "\n".join(f"  - {entry}" for entry in not_found)
            raise SystemExit(f"Cards not found in collection:\n{lines}")

        deck_id = insert_deck(conn, deck)
        insert_deck_cards(
            conn,
            deck_id,
            deck["cards"],
            lang_config.deck_nation_to_db,
            lang_config.locale_key,
        )
        conn.commit()

    logger.info("Deck '%s' imported (%d cards)", deck["name"], len(deck["cards"]))


def add_deck(
    filename: str,
    update: bool = False,
    replace: bool = False,
    db_path: str = DEFAULT_DB_PATH,
) -> None:
    """Add a deck from TXT file with exile card support and quantity check.

    Raises RuntimeError (not SystemExit) so the CLI can collect errors
    across multiple files and report a batch summary.

    Args:
        filename: Path to deck TXT file.
        update: If True, update collection quantities to match deck.
        replace: If True, replace existing deck with same name.
        db_path: SQLite database path.
    """
    lang_config = get_language_config()
    logger.info("Adding deck from file: %s", filename)

    try:
        deck = parse_deck_file(filename)
    except (FileNotFoundError, ValueError) as e:
        raise RuntimeError(f"Failed to parse deck file: {e}") from e

    with get_connection(db_path) as conn:
        initialize_schema(conn)

        existing = find_deck_by_name(conn, deck["name"])
        if existing:
            if not replace:
                if existing.get("deck_code") != deck["deck_code"]:
                    raise RuntimeError(
                        f"Deck '{deck['name']}' exists with different code. Use --replace."
                    )
                raise RuntimeError(
                    f"Deck '{deck['name']}' already exists (id={existing['deck_id']})"
                )
            delete_deck(conn, existing["deck_id"])
            conn.commit()

        # Resolve card IDs with exile fallback
        not_found = []
        resolved: list[tuple[DeckCardEntry, str]] = []
        for card in deck["cards"]:
            faction = lang_config.deck_nation_to_db.get(card["nation"], card["nation"])
            card_id = find_card_id(conn, faction, card["name"], lang_config.locale_key)
            if card_id is None:
                card_id = find_card_id_by_exile(conn, faction, card["name"], lang_config.locale_key)
            if card_id is None:
                not_found.append(f"{faction} / {card['name']}")
            else:
                resolved.append((card, card_id))

        if not_found:
            lines = "\n".join(f"  - {entry}" for entry in not_found)
            raise RuntimeError(f"Cards not found in collection:\n{lines}")

        # Quantity check
        mismatches: list[tuple[DeckCardEntry, str, int, int]] = []
        for card, card_id in resolved:
            faction = lang_config.deck_nation_to_db.get(card["nation"], card["nation"])
            collection_qty = get_card_quantity_by_id(conn, card_id)
            if card["quantity"] != collection_qty:
                mismatches.append((card, card_id, card["quantity"], collection_qty))

        if mismatches and not update:
            lines = "\n".join(
                f"  - {lang_config.deck_nation_to_db.get(c['nation'], c['nation'])} / {c['name']}:"
                f" deck={deck_qty}, collection={col_qty}"
                for c, _, deck_qty, col_qty in mismatches
            )
            raise RuntimeError(
                f"Card quantity mismatch:\n{lines}\n"
                "Re-run with --update (-u) to update collection quantities."
            )

        deck_id = insert_deck(conn, deck)
        insert_deck_cards(
            conn,
            deck_id,
            deck["cards"],
            lang_config.deck_nation_to_db,
            lang_config.locale_key,
            use_exile_fallback=True,
        )
        conn.commit()

        if update and mismatches:
            for _, card_id, deck_qty, _ in mismatches:
                update_card_quantity_by_id(conn, card_id, deck_qty)
            conn.commit()

    logger.info("Deck '%s' added (%d cards)", deck["name"], len(deck["cards"]))


def _select_deck(conn: sqlite3.Connection) -> dict:
    """Interactively select a deck from the database.

    Args:
        conn: SQLite connection instance.

    Returns:
        Selected deck metadata dict.
    """
    decks = fetch_all_decks(conn)
    if not decks:
        raise SystemExit("No decks in database. Run 'kards deck import' first.")

    print("Available decks:")
    for i, deck in enumerate(decks, 1):
        print(f"  {i}. {deck['name']}")

    try:
        choice = int(input("Enter deck number: "))
    except (ValueError, EOFError) as e:
        raise SystemExit(f"Invalid input: expected a deck number (1-{len(decks)})") from e

    if choice < 1 or choice > len(decks):
        raise SystemExit(f"Invalid choice: {choice}. Enter a number from 1 to {len(decks)}")

    return decks[choice - 1]


def remove_deck(db_path: str = DEFAULT_DB_PATH) -> None:
    """Interactively select and delete a deck from the database.

    Enter 0 to delete all decks. Prompts for confirmation before deleting.
    Does not affect the cards table.

    Args:
        db_path: SQLite database path.
    """
    with get_connection(db_path) as conn:
        initialize_schema(conn)
        decks = fetch_all_decks(conn)
        if not decks:
            raise SystemExit("No decks in database. Run 'kards deck import' first.")

        print("Available decks:")
        print("  0. Delete ALL decks")
        for i, deck in enumerate(decks, 1):
            print(f"  {i}. {deck['name']}")

        try:
            choice = int(input("Enter deck number (0 to delete all): "))
        except (ValueError, EOFError) as e:
            raise SystemExit(f"Invalid input: expected a number (0-{len(decks)})") from e

        if choice < 0 or choice > len(decks):
            raise SystemExit(f"Invalid choice: {choice}. Enter a number from 0 to {len(decks)}")

        if choice == 0:
            print(f"\nAll {len(decks)} deck(s) will be deleted.")
            try:
                confirm = input("Confirm deletion? [y/N]: ").strip().lower()
            except EOFError:
                raise SystemExit("Aborted.")
            if confirm != "y":
                raise SystemExit("Aborted.")
            count = delete_all_decks(conn)
            conn.commit()

        else:
            deck = decks[choice - 1]
            print(f"\nDeck to delete: {deck['name']}")
            try:
                confirm = input("Confirm deletion? [y/N]: ").strip().lower()
            except EOFError:
                raise SystemExit("Aborted.")
            if confirm != "y":
                raise SystemExit("Aborted.")
            delete_deck(conn, deck["deck_id"])
            conn.commit()

    if choice == 0:
        logger.info("All decks deleted (%d).", count)
    else:
        logger.info("Deck '%s' deleted.", deck["name"])


def export_deck(
    fmt: str | None,
    filename: str,
    db_path: str = DEFAULT_DB_PATH,
) -> None:
    """Export a deck to XLSX sheet or JSON file.

    Args:
        fmt: Export format ('json' or None for xlsx).
        filename: Output file path.
        db_path: SQLite database path.
    """
    lang_config = get_language_config()

    with get_connection(db_path) as conn:
        initialize_schema(conn)
        deck_meta = _select_deck(conn)
        deck_cards = fetch_deck_cards(conn, deck_meta["deck_id"])

    if not deck_cards:
        raise SystemExit("Deck has no cards")

    if fmt == "json":
        export_deck_to_json(deck_meta, deck_cards, filename, lang_config)
    else:
        wb = load_workbook(filename)
        add_deck_sheet(
            wb,
            deck_meta,
            deck_cards,
            lang_config.deck_headers,
            lang_config.deck_metadata_labels,
            lang_config.deck_nation_to_db,
            lang_config.nation_display_names,
            lang_config,
        )
        wb.save(filename)

    logger.info("Deck exported: %s", filename)


def add_match(db_path: str = DEFAULT_DB_PATH) -> None:
    """Interactively record match results for a deck.

    Args:
        db_path: SQLite database path.
    """
    lang_config = get_language_config()
    valid_factions = set(lang_config.faction_names.keys())

    with get_connection(db_path) as conn:
        initialize_schema(conn)
        deck = _select_deck(conn)
        count = 0

        while True:
            result = input("Result (win/loss): ").strip().lower()
            if result not in ("win", "loss"):
                print("Invalid result. Enter 'win' or 'loss'.")
                continue

            opponent_major = input("Opponent major power: ").strip()
            if opponent_major not in valid_factions:
                print(f"Invalid faction. Choose from: {', '.join(sorted(valid_factions))}")
                continue

            opponent_ally = input("Opponent ally: ").strip()
            if opponent_ally not in valid_factions:
                print(f"Invalid faction. Choose from: {', '.join(sorted(valid_factions))}")
                continue

            insert_match(conn, deck["deck_id"], result, opponent_major, opponent_ally)
            conn.commit()
            count += 1

            again = input("Add another match? [y/N]: ").strip().lower()
            if again != "y":
                break

    logger.info("%d match(es) recorded for deck '%s'.", count, deck["name"])


def analyze_deck(depth: str | None = None, db_path: str = DEFAULT_DB_PATH) -> None:
    """Analyze a deck using LLM advisor.

    Args:
        depth: Analysis depth override (concise/standard/detailed).
        db_path: SQLite database path.
    """
    from dotenv import load_dotenv

    lang_config = get_language_config()
    advisor_config = get_advisor_config()
    effective_depth = depth or advisor_config.depth

    with get_connection(db_path) as conn:
        initialize_schema(conn)
        deck = _select_deck(conn)
        cards = fetch_deck_cards(conn, deck["deck_id"])
        if not cards:
            raise SystemExit("Deck has no cards")
        stats = compute_deck_stats(conn, deck["deck_id"])

    system_prompt, user_prompt = build_analysis_context(
        deck, cards, stats, effective_depth, lang_config=lang_config,
    )
    load_dotenv()
    response = get_llm_response(system_prompt, user_prompt, advisor_config)
    print(response)
