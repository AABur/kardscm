"""Deck export helpers (XLSX sheet, JSON) for card decks."""

from __future__ import annotations

import json
import logging

from openpyxl.utils import get_column_letter
from openpyxl.workbook import Workbook as WorkbookType

from kardscm.config import LanguageConfig
from kardscm.constants import DECK_COLUMN_WIDTHS, DECK_NATION_TO_DB
from kardscm.export.styles import BOLD_FONT, HEADER_ALIGNMENT_NO_WRAP, HEADER_FILL, HEADER_FONT
from kardscm.helpers import extract_locale

logger = logging.getLogger(__name__)


def add_deck_sheet(
    wb: WorkbookType,
    deck_meta: dict,
    deck_cards: list[dict],
    deck_headers: list[str],
    metadata_labels: list[str],
    nation_display_names: dict[str, str],
    lang_config: LanguageConfig,
) -> None:
    """Add a deck sheet to an existing workbook.

    Args:
        wb: openpyxl Workbook instance.
        deck_meta: Deck metadata dict.
        deck_cards: List of card dicts with deck_quantity and deck_cost.
        deck_headers: Headers for deck card table.
        metadata_labels: Labels for deck metadata section.
        nation_display_names: Display names for nation sections.
        lang_config: Language configuration for translating card fields.
    """
    sheet_name = deck_meta["name"][:31]
    ws = wb.create_sheet(title=sheet_name)

    major = deck_meta.get("major_power", "")
    ally = deck_meta.get("ally", "")
    meta_values = [
        deck_meta.get("name", ""),
        lang_config.faction_names.get(DECK_NATION_TO_DB.get(major, major), major),
        lang_config.faction_names.get(DECK_NATION_TO_DB.get(ally, ally), ally),
        deck_meta.get("hq", ""),
        deck_meta.get("deck_code", ""),
    ]
    for row_idx, (label, value) in enumerate(zip(metadata_labels, meta_values), 1):
        ws.cell(row=row_idx, column=1, value=label).font = BOLD_FONT
        ws.cell(row=row_idx, column=2, value=value or "")

    header_row = 7
    for col_idx, header in enumerate(deck_headers, 1):
        cell = ws.cell(row=header_row, column=col_idx, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = HEADER_ALIGNMENT_NO_WRAP

    current_row = header_row + 1
    cards_by_faction: dict[str, list[dict]] = {}
    for card in deck_cards:
        faction = card.get("faction", "")
        cards_by_faction.setdefault(faction, []).append(card)

    locale_key = lang_config.locale_key
    for faction, faction_cards in cards_by_faction.items():
        faction_lower = faction.lower()
        display_name = nation_display_names.get(faction_lower, faction)
        ws.cell(row=current_row, column=1, value=display_name).font = BOLD_FONT
        current_row += 1

        for card in faction_cards:
            title_raw = card.get("title", "")
            title = extract_locale(title_raw, locale_key, default=str(title_raw), en_fallback=True)

            type_api = card.get("type", "")
            type_name = lang_config.type_names.get(type_api, type_api)

            ws.cell(row=current_row, column=1, value=title)
            ws.cell(row=current_row, column=2, value=type_name)
            ws.cell(row=current_row, column=3, value=card.get("deck_quantity"))
            ws.cell(row=current_row, column=4, value=card.get("deck_cost"))
            ws.cell(row=current_row, column=5, value=card.get("attack"))
            ws.cell(row=current_row, column=6, value=card.get("defense"))
            current_row += 1

    for col_idx, width in enumerate(DECK_COLUMN_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    logger.info("Deck sheet '%s' added (%d cards)", sheet_name, len(deck_cards))


def export_deck_to_json(
    deck_meta: dict,
    deck_cards: list[dict],
    filename: str,
    lang_config: LanguageConfig,
) -> None:
    """Export a deck to JSON format.

    Args:
        deck_meta: Deck metadata dict.
        deck_cards: List of card dicts with deck_quantity and deck_cost.
        filename: Output file path.
        lang_config: Language configuration for translating card fields.
    """
    locale_key = lang_config.locale_key

    def extract_title(card: dict) -> str:
        title_raw = card.get("title", "")
        return extract_locale(title_raw, locale_key, default=str(title_raw), en_fallback=True)

    output = {
        "deck": {
            "name": deck_meta.get("name", ""),
            "major_power": deck_meta.get("major_power", ""),
            "ally": deck_meta.get("ally"),
            "hq": deck_meta.get("hq"),
        },
        "cards": [
            {
                "faction": lang_config.faction_names.get(
                    card.get("faction", ""), card.get("faction", "")
                ),
                "title": extract_title(card),
                "type": lang_config.type_names.get(card.get("type", ""), card.get("type", "")),
                "rarity": lang_config.rarity_names.get(
                    card.get("rarity", ""), card.get("rarity", "")
                ),
                "set": lang_config.set_names.get(card.get("set", ""), card.get("set", "")),
                "kredits": card.get("kredits"),
                "attack": card.get("attack"),
                "defense": card.get("defense"),
                "text": card.get("text", ""),
                "deck_quantity": card.get("deck_quantity"),
                "deck_cost": card.get("deck_cost"),
            }
            for card in deck_cards
        ],
    }

    with open(filename, "w", encoding="utf-8") as f:
        json.dump(output, f, ensure_ascii=False, indent=2)

    logger.info("Deck JSON exported: %s (%d cards)", filename, len(deck_cards))
