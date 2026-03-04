"""Export helpers for card collections."""

from __future__ import annotations

import csv
import json
import logging

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.workbook import Workbook as WorkbookType

from kardscm.config import LanguageConfig
from kardscm.constants import DECK_COLUMN_WIDTHS, EXPORT_FIELD_NAMES
from kardscm.helpers import sanitize_text

logger = logging.getLogger(__name__)


def translate_card_for_export(card: dict, lang_config: LanguageConfig) -> dict:
    """Translate a raw DB card dict to a localized export dict.

    Extracts localized title and text via locale_key, translates
    faction/type/rarity/set via static mappings, formats attributes.

    Args:
        card: Raw card dict from database.
        lang_config: Language configuration.

    Returns:
        Dict with export field names as keys.
    """
    locale_key = lang_config.locale_key

    # Extract localized title
    title_raw = card.get("title", "")
    try:
        title_dict = json.loads(title_raw) if title_raw else {}
    except (json.JSONDecodeError, TypeError):
        title_dict = {}
    if isinstance(title_dict, dict):
        title = title_dict.get(locale_key, title_dict.get("en-EN", title_raw))
    else:
        title = str(title_raw)

    # Extract localized text
    text_raw = card.get("text", "")
    try:
        text_dict = json.loads(text_raw) if text_raw else {}
    except (json.JSONDecodeError, TypeError):
        text_dict = {}
    if isinstance(text_dict, dict):
        text = text_dict.get(locale_key, text_dict.get("en-EN", ""))
    else:
        text = str(text_raw) if text_raw else ""

    # Translate faction
    faction_api = card.get("faction", "")
    faction = lang_config.faction_names.get(faction_api, faction_api)

    # Translate type
    type_api = card.get("type", "")
    type_name = lang_config.type_names.get(type_api, type_api)

    # Translate rarity
    rarity_api = card.get("rarity", "")
    rarity = lang_config.rarity_names.get(rarity_api, rarity_api)

    # Translate set
    set_api = card.get("set", "")
    set_name = lang_config.set_names.get(set_api, set_api)

    # Format attributes
    attributes_raw = card.get("attributes", "[]")
    try:
        attributes_list = json.loads(attributes_raw) if attributes_raw else []
    except (json.JSONDecodeError, TypeError):
        attributes_list = []
    abilities = ", ".join(
        lang_config.ability_names.get(a, a)
        for a in attributes_list
        if a in lang_config.ability_names
    )

    return {
        "faction": sanitize_text(faction),
        "title": sanitize_text(title),
        "type": sanitize_text(type_name),
        "rarity": sanitize_text(rarity),
        "attributes": abilities,
        "set": sanitize_text(set_name),
        "quantity": card.get("quantity", 0),
        "kredits": card.get("kredits", 0),
        "attack": card.get("attack"),
        "defense": card.get("defense"),
        "text": sanitize_text(str(text)) if text else "",
    }


def export_to_xlsx(
    cards: list[dict],
    filename: str,
    headers: list[str],
    sheet_name: str = "Collection",
) -> None:
    """Export cards to Excel format with formatting.

    Args:
        cards: List of card dictionaries (already translated for export).
        filename: Output filename.
        headers: Display headers for columns.
        sheet_name: Name of the worksheet.
    """
    logger.info("Exporting %s cards to Excel: %s", len(cards), filename)

    wb = Workbook()
    ws = wb.active
    if not ws:
        msg = "Failed to create worksheet"
        raise RuntimeError(msg)
    ws.title = sheet_name

    header_row = list(headers)
    ws.append(header_row)

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    numeric_fields = {"quantity", "kredits", "attack", "defense"}
    for card in cards:
        row = [
            card.get(field) if field in numeric_fields else card.get(field, "")
            for field in EXPORT_FIELD_NAMES
        ]
        ws.append(row)

    column_widths = [15, 35, 18, 15, 12, 20, 10, 10, 8, 8, 60]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(header_row))}{len(cards) + 1}"

    wb.save(filename)
    logger.info("Excel file created: %s (%s cards)", filename, len(cards))


def export_to_csv(
    cards: list[dict],
    filename: str,
    headers: list[str],
) -> None:
    """Export cards to CSV format.

    Uses UTF-8 with BOM for Excel compatibility on Windows.

    Args:
        cards: List of card dictionaries (already translated for export).
        filename: Output filename.
        headers: Display headers for columns.
    """
    logger.info("Exporting %s cards to CSV: %s", len(cards), filename)

    header_row = list(headers)

    with open(filename, "w", newline="", encoding="utf-8-sig") as csvfile:
        writer = csv.writer(csvfile)
        writer.writerow(header_row)

        for card in cards:
            row = [card.get(field, "") for field in EXPORT_FIELD_NAMES]
            writer.writerow(row)

    logger.info("CSV file created: %s (%s cards)", filename, len(cards))


def export_to_json(
    cards: list[dict],
    filename: str,
    language: str,
    language_name: str,
) -> None:
    """Export cards to JSON format with metadata.

    Args:
        cards: List of card dictionaries (already translated for export).
        filename: Output filename.
        language: Language code for metadata.
        language_name: Language name for metadata.
    """
    logger.info("Exporting %s cards to JSON: %s", len(cards), filename)

    output_data = {
        "metadata": {
            "language": language,
            "language_name": language_name,
            "total_cards": len(cards),
        },
        "cards": cards,
    }

    with open(filename, "w", encoding="utf-8") as jsonfile:
        json.dump(output_data, jsonfile, ensure_ascii=False, indent=2)

    logger.info("JSON file created: %s (%s cards)", filename, len(cards))


def add_deck_sheet(
    wb: WorkbookType,
    deck_meta: dict,
    deck_cards: list[dict],
    deck_headers: list[str],
    metadata_labels: list[str],
    deck_nation_to_db: dict[str, str],
    nation_display_names: dict[str, str],
    lang_config: LanguageConfig,
) -> None:
    """Add a deck sheet to an existing workbook.

    Args:
        wb: openpyxl Workbook instance.
        deck_meta: Deck metadata dict.
        deck_cards: List of card dicts with deck_quantity and deck_cost.
        deck_headers: Headers for deck card table.
        metadata_labels: Labels for deck metadata section.
        deck_nation_to_db: Mapping from deck nation keys to API faction names.
        nation_display_names: Display names for nation sections.
        lang_config: Language configuration for translating card fields.
    """
    sheet_name = deck_meta["name"][:31]
    ws = wb.create_sheet(title=sheet_name)

    bold_font = Font(bold=True)

    major = deck_meta.get("major_power", "")
    ally = deck_meta.get("ally", "")
    meta_values = [
        deck_meta.get("name", ""),
        lang_config.faction_names.get(deck_nation_to_db.get(major, major), major),
        lang_config.faction_names.get(deck_nation_to_db.get(ally, ally), ally),
        deck_meta.get("hq", ""),
        deck_meta.get("deck_code", ""),
    ]
    for row_idx, (label, value) in enumerate(zip(metadata_labels, meta_values), 1):
        ws.cell(row=row_idx, column=1, value=label).font = bold_font
        ws.cell(row=row_idx, column=2, value=value or "")

    header_row = 7
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    for col_idx, header in enumerate(deck_headers, 1):
        cell = ws.cell(row=header_row, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    current_row = header_row + 1
    cards_by_faction: dict[str, list[dict]] = {}
    for card in deck_cards:
        faction = card.get("faction", "")
        cards_by_faction.setdefault(faction, []).append(card)

    locale_key = lang_config.locale_key
    for faction, faction_cards in cards_by_faction.items():
        faction_lower = faction.lower()
        display_name = nation_display_names.get(faction_lower, faction)
        ws.cell(row=current_row, column=1, value=display_name).font = bold_font
        current_row += 1

        for card in faction_cards:
            # Extract localized title
            title_raw = card.get("title", "")
            try:
                title_dict = json.loads(title_raw) if title_raw else {}
            except (json.JSONDecodeError, TypeError):
                title_dict = {}
            if isinstance(title_dict, dict):
                title = title_dict.get(locale_key, title_dict.get("en-EN", title_raw))
            else:
                title = str(title_raw)

            type_api = card.get("type", "")
            type_name = lang_config.type_names.get(type_api, type_api)

            ws.cell(row=current_row, column=1, value=title)
            ws.cell(row=current_row, column=2, value=type_name)
            ws.cell(row=current_row, column=3, value=card.get("deck_quantity"))
            ws.cell(row=current_row, column=4, value=card.get("deck_cost"))
            ws.cell(row=current_row, column=5, value=card.get("attack"))
            ws.cell(row=current_row, column=6, value=card.get("defense"))
            current_row += 1

    for col_idx, width in enumerate(DECK_COLUMN_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    logger.info("Deck sheet '%s' added (%d cards)", sheet_name, len(deck_cards))


def export_deck_to_json(
    deck_meta: dict,
    deck_cards: list[dict],
    filename: str,
    lang_config: LanguageConfig,
) -> None:
    """Export a deck to JSON format.

    Args:
        deck_meta: Deck metadata dict.
        deck_cards: List of card dicts with deck_quantity and deck_cost.
        filename: Output file path.
        lang_config: Language configuration for translating card fields.
    """
    locale_key = lang_config.locale_key

    def extract_title(card: dict) -> str:
        title_raw = card.get("title", "")
        try:
            title_dict = json.loads(title_raw) if title_raw else {}
        except (json.JSONDecodeError, TypeError):
            title_dict = {}
        if isinstance(title_dict, dict):
            return str(title_dict.get(locale_key, title_dict.get("en-EN", title_raw)))
        return str(title_raw)

    output = {
        "deck": {
            "name": deck_meta.get("name", ""),
            "major_power": deck_meta.get("major_power", ""),
            "ally": deck_meta.get("ally"),
            "hq": deck_meta.get("hq"),
        },
        "cards": [
            {
                "faction": lang_config.faction_names.get(
                    card.get("faction", ""), card.get("faction", "")
                ),
                "title": extract_title(card),
                "type": lang_config.type_names.get(card.get("type", ""), card.get("type", "")),
                "rarity": lang_config.rarity_names.get(
                    card.get("rarity", ""), card.get("rarity", "")
                ),
                "set": lang_config.set_names.get(card.get("set", ""), card.get("set", "")),
                "kredits": card.get("kredits"),
                "attack": card.get("attack"),
                "defense": card.get("defense"),
                "text": card.get("text", ""),
                "deck_quantity": card.get("deck_quantity"),
                "deck_cost": card.get("deck_cost"),
            }
            for card in deck_cards
        ],
    }

    with open(filename, "w", encoding="utf-8") as f:
        json.dump(output, f, ensure_ascii=False, indent=2)

    logger.info("Deck JSON exported: %s (%d cards)", filename, len(deck_cards))
