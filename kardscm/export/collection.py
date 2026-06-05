"""Collection export helpers (XLSX, JSON) for card collections.

XLSX export is an exact replica of the 12-column web collection table
(localized values). JSON export is a raw API-shape dump for downstream
LLM use: raw codes, locale objects, ability arrays, plus quantity — not
affected by ``--lang``.
"""

from __future__ import annotations

import json
import logging

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from kardscm.config import LanguageConfig
from kardscm.constants import (
    COLLECTION_COLUMN_WIDTHS,
    COLLECTION_TABLE_FIELDS,
    KNOWN_ABILITIES,
    KNOWN_EXTRA_ABILITIES,
)
from kardscm.export.styles import HEADER_ALIGNMENT, HEADER_FILL, HEADER_FONT
from kardscm.helpers import extract_locale, sanitize_text

logger = logging.getLogger(__name__)

# Fields written to the XLSX as numbers rather than strings.
_NUMERIC_FIELDS = frozenset({"quantity", "kredits", "operationCost", "attack", "defense"})

# Raw API fields copied verbatim from the DB row into the JSON export. The
# remaining JSON keys (title/text/can_create, attributes/extra_abilities,
# quantity) are computed in card_to_api_dict.
_API_PASSTHROUGH_FIELDS = (
    "cardId",
    "importId",
    "faction",
    "type",
    "rarity",
    "set",
    "kredits",
    "attack",
    "defense",
    "operationCost",
    "reserved",
    "image",
    "imageUrl",
    "thumbUrl",
    "exile",
)


def translate_card_for_export(card: dict, lang_config: LanguageConfig) -> dict:
    """Translate a raw DB card dict to a localized export dict.

    Extracts localized title and text via locale_key, translates
    faction/type/rarity/set via static mappings, and formats both the
    API abilities and the curated extra abilities into localized,
    comma-joined strings. The returned dict carries all 12 web-table
    fields plus ``text`` (kept for reuse; unused by the XLSX export).

    Args:
        card: Raw card dict from database.
        lang_config: Language configuration.

    Returns:
        Dict with export field names as keys.
    """
    locale_key = lang_config.locale_key

    title_raw = card.get("title", "")
    title = extract_locale(title_raw, locale_key, default=str(title_raw), en_fallback=True)

    text_raw = card.get("text", "")
    text = extract_locale(text_raw, locale_key, en_fallback=True)

    # Translate faction
    faction_api = card.get("faction", "")
    faction = lang_config.faction_names.get(faction_api, faction_api)

    # Translate type
    type_api = card.get("type", "")
    type_name = lang_config.type_names.get(type_api, type_api)

    # Translate rarity
    rarity_api = card.get("rarity", "")
    rarity = lang_config.rarity_names.get(rarity_api, rarity_api)

    # Translate set
    set_api = card.get("set", "")
    set_name = lang_config.set_names.get(set_api, set_api)

    # Format abilities from binary columns
    abilities = ", ".join(
        lang_config.ability_names.get(a, a) for a in KNOWN_ABILITIES if card.get(f"ability_{a}", 0)
    )

    # Format curated extra abilities from binary columns
    extra_attributes = sanitize_text(
        ", ".join(
            lang_config.extra_ability_names.get(a, a)
            for a in KNOWN_EXTRA_ABILITIES
            if card.get(f"extra_ability_{a}", 0)
        )
    )

    return {
        "faction": sanitize_text(faction),
        "title": sanitize_text(title),
        "type": sanitize_text(type_name),
        "rarity": sanitize_text(rarity),
        "attributes": abilities,
        "extra_attributes": extra_attributes,
        "set": sanitize_text(set_name),
        "quantity": card.get("quantity", 0),
        "kredits": card.get("kredits", 0),
        "operationCost": card.get("operationCost"),
        "attack": card.get("attack"),
        "defense": card.get("defense"),
        "text": sanitize_text(str(text)) if text else "",
    }


def build_collection_headers(lang_config: LanguageConfig) -> list[str]:
    """Build the 12 collection XLSX headers in web-table order.

    Resolves each ``COLLECTION_TABLE_FIELDS`` column to its localized
    header: most come from ``export_headers``; the two web-only columns
    (extra abilities, cost) come from ``ui_strings``. Driving the row off
    ``COLLECTION_TABLE_FIELDS`` keeps the column order single-sourced and
    the XLSX header row identical to the rendered web collection table.

    Args:
        lang_config: Active language configuration.

    Returns:
        Twelve header strings in web-table order.
    """
    h = lang_config.export_headers
    ui = lang_config.ui_strings
    header_by_field = {
        "faction": h[0],
        "title": h[1],
        "type": h[2],
        "rarity": h[3],
        "attributes": h[4],
        "extra_attributes": ui["filter_extra_abilities"],
        "set": h[5],
        "quantity": h[6],
        "kredits": h[7],
        "operationCost": ui["col_cost"],
        "attack": h[8],
        "defense": h[9],
    }
    return [header_by_field[field] for field in COLLECTION_TABLE_FIELDS]


def export_to_xlsx(
    cards: list[dict],
    filename: str,
    headers: list[str],
    sheet_name: str = "Collection",
) -> None:
    """Export cards to Excel format with formatting.

    Args:
        cards: List of card dictionaries (already translated for export).
        filename: Output filename.
        headers: Display headers for columns.
        sheet_name: Name of the worksheet.
    """
    logger.info("Exporting %s cards to Excel: %s", len(cards), filename)

    wb = Workbook()
    ws = wb.active
    if not ws:
        msg = "Failed to create worksheet"
        raise RuntimeError(msg)
    ws.title = sheet_name

    header_row = list(headers)
    ws.append(header_row)

    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = HEADER_ALIGNMENT

    for card in cards:
        row = [
            card.get(field) if field in _NUMERIC_FIELDS else card.get(field, "")
            for field in COLLECTION_TABLE_FIELDS
        ]
        ws.append(row)

    for i, field in enumerate(COLLECTION_TABLE_FIELDS, 1):
        ws.column_dimensions[get_column_letter(i)].width = COLLECTION_COLUMN_WIDTHS[field]

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(header_row))}{len(cards) + 1}"

    wb.save(filename)
    logger.info("Excel file created: %s (%s cards)", filename, len(cards))


def _decode_json_field(value: object) -> object:
    """Decode a field stored as a JSON string back to its object form.

    Card ``title``/``text``/``can_create`` are persisted as JSON strings.
    Returns the parsed value on success, or the original value unchanged
    when it is not a string or cannot be decoded.
    """
    if not isinstance(value, str):
        return value
    try:
        return json.loads(value)
    except json.JSONDecodeError:
        return value


def card_to_api_dict(raw_card: dict) -> dict:
    """Build a raw API-shape dict from a raw database row.

    Reconstructs the downstream-LLM JSON shape: raw codes for
    faction/type/rarity/set, ``title``/``text``/``can_create`` parsed
    back to locale objects, ``attributes``/``extra_abilities`` as arrays
    of raw keys, plus per-card quantity and the remaining API fields.
    Performs no localization or text sanitization.

    Args:
        raw_card: Raw card dict from the database.

    Returns:
        Raw API-shape dict ready for JSON serialization.
    """
    result = {field: raw_card.get(field) for field in _API_PASSTHROUGH_FIELDS}
    result["title"] = _decode_json_field(raw_card.get("title"))
    result["text"] = _decode_json_field(raw_card.get("text"))
    result["can_create"] = _decode_json_field(raw_card.get("can_create"))
    result["attributes"] = [a for a in KNOWN_ABILITIES if raw_card.get(f"ability_{a}")]
    result["extra_abilities"] = [
        a for a in KNOWN_EXTRA_ABILITIES if raw_card.get(f"extra_ability_{a}")
    ]
    result["quantity"] = raw_card.get("quantity", 0)
    return result


def export_to_json(raw_cards: list[dict], filename: str) -> None:
    """Export raw cards to JSON in raw API shape for downstream LLM use.

    The payload is not localized and is unaffected by ``--lang``.

    Args:
        raw_cards: Raw card dicts from the database.
        filename: Output filename.
    """
    logger.info("Exporting %s cards to JSON: %s", len(raw_cards), filename)

    output_data = {
        "metadata": {"total_cards": len(raw_cards)},
        "cards": [card_to_api_dict(card) for card in raw_cards],
    }

    with open(filename, "w", encoding="utf-8") as jsonfile:
        json.dump(output_data, jsonfile, ensure_ascii=False, indent=2)

    logger.info("JSON file created: %s (%s cards)", filename, len(raw_cards))
