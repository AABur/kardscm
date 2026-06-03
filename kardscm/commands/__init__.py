"""Business logic for CLI commands."""

from __future__ import annotations

import json
import logging
import shutil
import sqlite3
from pathlib import Path

import typer
from openpyxl import load_workbook

from kardscm.commands.utils import (
    _default_diff_report_path,
    _emit_locale_warnings,
    _safe_timestamp,
    _utc_timestamp,
)
from kardscm.commands.validation import validate_file as validate_file
from kardscm.config import LanguageConfig, get_language_config
from kardscm.constants import DECK_NATION_TO_DB, DEFAULT_DB_PATH
from kardscm.diff import (
    compute_diff,
    format_console_report,
    format_markdown_report,
    is_empty,
)
from kardscm.export import (
    add_deck_sheet,
    export_deck_to_json,
    export_to_csv,
    export_to_json,
    export_to_xlsx,
    translate_card_for_export,
)
from kardscm.helpers import parse_int
from kardscm.importing import parse_deck_file
from kardscm.models import CardDict, DeckCardEntry, DiffReport
from kardscm.scraping import baseline, fetcher, probe, scrape_cards
from kardscm.storage import (
    apply_extra_abilities_seed,
    delete_all_decks,
    delete_cards,
    delete_deck,
    fetch_all_decks,
    fetch_cards,
    fetch_deck_cards,
    find_card_id,
    find_card_id_by_exile,
    find_deck_by_name,
    get_card_quantity_by_id,
    get_connection,
    initialize_schema,
    insert_deck,
    insert_deck_cards,
    set_metadata,
    update_card_quantity_by_id,
    update_quantity,
    upsert_cards,
)

logger = logging.getLogger(__name__)


_APPROVAL_CATEGORIES = (
    ("new", "Apply"),
    ("changed", "Apply"),
    ("reserved_in", "Apply"),
    ("reserved_out", "Apply"),
    ("removed", "Delete"),
)


def _approve_all_categories(report: DiffReport, lang_config: LanguageConfig) -> bool:
    """Prompt y/N for each non-empty category. Any 'no' returns False."""
    headers = lang_config.diff_headers
    for key, verb in _APPROVAL_CATEGORIES:
        items = report[key]  # type: ignore[literal-required]
        if not items:
            continue
        label = headers.get(key, key)
        if not typer.confirm(f"{verb} {len(items)} '{label}'?", default=True):
            return False
    return True


def _write_diff_report(
    path: Path,
    report: DiffReport,
    lang_config: LanguageConfig,
    timestamp: str,
) -> None:
    content = format_markdown_report(report, lang_config, timestamp)
    path.write_text(content, encoding="utf-8")


def _read_xlsx_quantities(
    filename: str,
    lang_config: LanguageConfig,
) -> list[tuple[str, str, int | None]]:
    """Read faction, title, quantity from XLSX file.

    Column names are determined by the active LanguageConfig.
    Returns (faction_display, localized_title, quantity) tuples.

    Args:
        filename: Path to XLSX file.
        lang_config: Active language configuration (drives column headers).

    Returns:
        List of (faction, title, quantity) tuples.

    Raises:
        FileNotFoundError: If file doesn't exist.
        ValueError: If required columns missing.
    """
    if not Path(filename).exists():
        raise FileNotFoundError(f"File not found: {filename}")

    headers_list = lang_config.export_headers
    header_map = {
        headers_list[0]: "faction",
        headers_list[1]: "title",
        headers_list[6]: "quantity",
    }

    wb = load_workbook(filename)
    ws = wb.active
    if not ws:
        raise ValueError("No active worksheet found")

    headers = {
        header_map[cell.value]: col_idx
        for col_idx, cell in enumerate(ws[1], 1)
        if cell.value in header_map
    }
    if len(headers) != len(header_map):
        expected = ", ".join(header_map.keys())
        raise ValueError(f"Missing required columns: {expected}")

    results = []
    for row in ws.iter_rows(min_row=2, values_only=False):
        faction_cell = row[headers["faction"] - 1]
        title_cell = row[headers["title"] - 1]
        qty_cell = row[headers["quantity"] - 1]

        faction = faction_cell.value
        title = title_cell.value
        qty_val = qty_cell.value

        if not faction or not title:
            continue

        qty = parse_int(qty_val)
        results.append((str(faction).strip(), str(title).strip(), qty))

    return results


def fetch_and_compute_diff(
    db_path: str,
    lang_config: LanguageConfig,
) -> tuple[DiffReport, list[CardDict], str]:
    """Fetch fresh cards from the website and compute the DB diff.

    Pure read path: no DB writes, no console echo, no markdown report.
    Web flows call this to drive the preview modal; the CLI orchestrator
    composes it with `apply_sync_changes`.

    Args:
        db_path: SQLite database path.
        lang_config: Active language configuration.

    Returns:
        Tuple of (DiffReport, fetched cards, filesystem-safe UTC timestamp).
        The timestamp is generated once here so every report path the
        caller writes shares the same identifier.
    """
    new_cards = scrape_cards(language=lang_config.code, lang_config=lang_config)
    with get_connection(db_path) as conn:
        initialize_schema(conn, db_path)
        old_cards = fetch_cards(conn)
        report = compute_diff(old_cards, new_cards, lang_config.locale_key)
    return report, new_cards, _safe_timestamp()


def apply_sync_changes(
    db_path: str,
    new_cards: list[CardDict],
    report: DiffReport,
    lang_config: LanguageConfig,
    timestamp: str,
    diff_report_path: Path | None = None,
) -> Path | None:
    """Persist the sync result to the DB and update metadata.

    Args:
        db_path: SQLite database path.
        new_cards: Cards returned by `fetch_and_compute_diff`.
        report: Diff report bucketed by category.
        lang_config: Active language configuration.
        timestamp: Filesystem-safe UTC timestamp from `fetch_and_compute_diff`.
        diff_report_path: Optional override for the markdown report path.
            Defaults to `./sync-diff-<timestamp>.md` when a report is written.

    Returns:
        Path to the markdown report when a non-empty diff was applied,
        otherwise None (empty diff → metadata-only update).
    """
    with get_connection(db_path) as conn:
        initialize_schema(conn, db_path)
        if is_empty(report):
            set_metadata(conn, "last_sync", _utc_timestamp())
            set_metadata(conn, "language", lang_config.code)
            return None

        upsert_cards(conn, new_cards)
        if report["removed"]:
            delete_cards(conn, [r["cardId"] for r in report["removed"]])
        apply_extra_abilities_seed(conn)
        set_metadata(conn, "last_sync", _utc_timestamp())
        set_metadata(conn, "language", lang_config.code)

    report_path = diff_report_path or _default_diff_report_path()
    _write_diff_report(report_path, report, lang_config, timestamp)
    return report_path


def sync_collection(
    db_path: str = DEFAULT_DB_PATH,
    *,
    lang: str | None = None,
    diff_only: bool = False,
    yes: bool = False,
    diff_report_path: Path | None = None,
) -> None:
    """Synchronize local SQLite storage with the website.

    Computes a diff between the current DB state and the fresh API pull,
    prints it, and asks the user to bulk-approve each non-empty category.
    Any rejection aborts the sync — the DB is left untouched. The
    Markdown diff report is written whenever the diff is non-empty.

    Args:
        db_path: SQLite database path.
        lang: Active language code (e.g. "en", "ru"). Defaults to English.
        diff_only: If True, write the report and return without prompting
            or modifying the DB. Useful for previews and CI.
        yes: If True, auto-approve every category without prompting.
        diff_report_path: Override the default report path
            (`./sync-diff-<UTC-iso>.md`).
    """
    lang_config = get_language_config(lang)
    _emit_locale_warnings(lang_config)
    logger.info("Starting sync from website (language: %s)...", lang_config.name)

    report, new_cards, timestamp = fetch_and_compute_diff(db_path, lang_config)

    if is_empty(report):
        apply_sync_changes(db_path, new_cards, report, lang_config, timestamp)
        logger.info("No changes. %s cards in collection.", len(new_cards))
        return

    typer.echo(format_console_report(report, lang_config))
    report_path = diff_report_path or _default_diff_report_path()

    if diff_only:
        _write_diff_report(report_path, report, lang_config, timestamp)
        logger.info("Diff report written to %s. No DB changes.", report_path)
        return

    if not yes and not _approve_all_categories(report, lang_config):
        _write_diff_report(report_path, report, lang_config, timestamp)
        logger.info("Sync aborted by user. Diff report written to %s.", report_path)
        return

    written = apply_sync_changes(db_path, new_cards, report, lang_config, timestamp, report_path)
    logger.info("Sync completed. Stored %s cards. Report: %s", len(new_cards), written)


def baseline_init(*, lang: str | None = None) -> None:
    """Pull from live API and overwrite the committed baseline.

    For one-off use after cloning the repo or after intentional API changes.
    Always overwrites; use ``baseline_accept`` to promote a sync-generated
    observed snapshot instead.
    """
    lang_config = get_language_config(lang)
    logger.info("Fetching cards from API to rebuild baseline...")
    raw = fetcher.fetch_all_cards(probe.build_static_probe(language=lang_config.code))
    snapshot = baseline.build_snapshot(raw)
    baseline.save_baseline(snapshot)
    logger.info(
        "Baseline written to %s (%d cards, %d enum value sets).",
        baseline.BASELINE_PATH,
        snapshot["card_count"],
        len(snapshot["enum_values"]),
    )


_BASELINE_REQUIRED_KEYS = ("card_count", "node_keys", "json_keys", "enum_values")


def baseline_accept() -> None:
    """Promote the most recent ``sync-schema-observed-*.json`` to baseline.

    Looks for files matching the pattern in cwd, picks the lexicographically
    latest (timestamps are ISO-like and sort correctly), validates its
    structural shape (must be a dict with all required snapshot keys of
    the correct types), and copies it to the committed baseline location.
    """
    candidates = sorted(Path.cwd().glob("sync-schema-observed-*.json"))
    if not candidates:
        raise SystemExit(
            "No sync-schema-observed-*.json files found in current directory. "
            "Run `kardscm sync` first."
        )
    latest = candidates[-1]
    try:
        parsed = json.loads(latest.read_text(encoding="utf-8"))
    except (json.JSONDecodeError, OSError) as exc:
        raise SystemExit(f"Cannot parse {latest}: {exc}") from exc

    if not isinstance(parsed, dict):
        raise SystemExit(f"{latest} is not a snapshot object (got {type(parsed).__name__}).")
    missing = [k for k in _BASELINE_REQUIRED_KEYS if k not in parsed]
    if missing:
        raise SystemExit(f"{latest} is missing required keys: {', '.join(missing)}")
    if not isinstance(parsed["card_count"], int):
        raise SystemExit(f"{latest}: card_count must be an int")
    if not isinstance(parsed["node_keys"], list):
        raise SystemExit(f"{latest}: node_keys must be a list")
    if not isinstance(parsed["json_keys"], dict):
        raise SystemExit(f"{latest}: json_keys must be a dict")
    if not isinstance(parsed["enum_values"], dict):
        raise SystemExit(f"{latest}: enum_values must be a dict")

    shutil.copy2(latest, baseline.BASELINE_PATH)
    logger.info("Baseline updated from %s.", latest.name)


def export_collection(
    export_format: str,
    filename: str,
    db_path: str = DEFAULT_DB_PATH,
    *,
    lang: str | None = None,
) -> None:
    """Export cards from SQLite to the selected format.

    Args:
        export_format: Export format (csv, json, xlsx).
        filename: Output file path.
        db_path: SQLite database path.
        lang: Active language code (e.g. "en", "ru"). Defaults to English.
    """
    lang_config = get_language_config(lang)
    _emit_locale_warnings(lang_config)

    with get_connection(db_path) as conn:
        initialize_schema(conn, db_path)
        raw_cards = fetch_cards(conn)

    if not raw_cards:
        raise SystemExit("No cards in database. Run 'kards sync' first.")

    cards = [translate_card_for_export(card, lang_config) for card in raw_cards]

    if export_format == "xlsx":
        export_to_xlsx(
            cards,
            filename,
            lang_config.export_headers,
            lang_config.collection_sheet_name,
        )
    elif export_format == "csv":
        export_to_csv(cards, filename, lang_config.export_headers)
    elif export_format == "json":
        export_to_json(cards, filename, lang_config.code, lang_config.name)
    else:
        msg = f"Unsupported format: {export_format}"
        raise ValueError(msg)

    logger.info("Export completed: %s", filename)


def update_collection(
    filename: str,
    db_path: str = DEFAULT_DB_PATH,
    *,
    lang: str | None = None,
) -> None:
    """Update card quantities from XLSX file.

    Args:
        filename: XLSX file path.
        db_path: SQLite database path.
        lang: Active language code (e.g. "en", "ru"). Defaults to English.
    """
    logger.info("Starting update from file: %s", filename)
    lang_config = get_language_config(lang)
    _emit_locale_warnings(lang_config)

    try:
        updates = _read_xlsx_quantities(filename, lang_config)
    except (FileNotFoundError, ValueError) as e:
        raise SystemExit(f"Failed to read file: {e}") from e

    if not updates:
        logger.warning("No valid entries found in file")
        return

    # Reverse-map localized faction names to API faction names
    reverse_faction = {v: k for k, v in lang_config.faction_names.items()}

    mapped_updates = []
    for faction_display, title, qty in updates:
        faction_api = reverse_faction.get(faction_display, faction_display)
        mapped_updates.append((faction_api, title, qty))

    with get_connection(db_path) as conn:
        initialize_schema(conn, db_path)
        updated_count, not_found = update_quantity(conn, mapped_updates, lang_config.locale_key)

    logger.info(
        "Update completed: %d cards updated, %d not found",
        updated_count,
        len(not_found),
    )

    if not_found:
        for key in not_found:
            logger.warning("Card not found: %s", key)


def import_deck(
    filename: str,
    db_path: str = DEFAULT_DB_PATH,
    *,
    lang: str | None = None,
) -> None:
    """Import a deck from TXT file into the database.

    Args:
        filename: Path to deck TXT file.
        db_path: SQLite database path.
        lang: Active language code (e.g. "en", "ru"). Defaults to English.
    """
    lang_config = get_language_config(lang)
    _emit_locale_warnings(lang_config)
    logger.info("Importing deck from file: %s", filename)

    try:
        deck = parse_deck_file(filename)
    except (FileNotFoundError, ValueError) as e:
        raise SystemExit(f"Failed to parse deck file: {e}") from e

    with get_connection(db_path) as conn:
        initialize_schema(conn, db_path)

        existing = find_deck_by_name(conn, deck["name"])
        if existing:
            raise SystemExit(f"Deck '{deck['name']}' already exists (id={existing['deck_id']})")

        not_found = []
        for card in deck["cards"]:
            faction = DECK_NATION_TO_DB.get(card["nation"], card["nation"])
            card_id = find_card_id(conn, faction, card["name"], lang_config.locale_key)
            if card_id is None:
                not_found.append(f"{faction} / {card['name']}")

        if not_found:
            lines = "\n".join(f"  - {entry}" for entry in not_found)
            raise SystemExit(f"Cards not found in collection:\n{lines}")

        deck_id = insert_deck(conn, deck)
        insert_deck_cards(
            conn,
            deck_id,
            deck["cards"],
            lang_config.locale_key,
        )
        conn.commit()

    logger.info("Deck '%s' imported (%d cards)", deck["name"], len(deck["cards"]))


def add_deck(
    filename: str,
    update: bool = False,
    replace: bool = False,
    db_path: str = DEFAULT_DB_PATH,
    *,
    lang: str | None = None,
) -> None:
    """Add a deck from TXT file with exile card support and quantity check.

    Raises RuntimeError (not SystemExit) so the CLI can collect errors
    across multiple files and report a batch summary.

    Args:
        filename: Path to deck TXT file.
        update: If True, update collection quantities to match deck.
        replace: If True, replace existing deck with same name.
        db_path: SQLite database path.
        lang: Active language code (e.g. "en", "ru"). Defaults to English.
    """
    lang_config = get_language_config(lang)
    _emit_locale_warnings(lang_config)
    logger.info("Adding deck from file: %s", filename)

    try:
        deck = parse_deck_file(filename)
    except (FileNotFoundError, ValueError) as e:
        raise RuntimeError(f"Failed to parse deck file: {e}") from e

    with get_connection(db_path) as conn:
        initialize_schema(conn, db_path)

        existing = find_deck_by_name(conn, deck["name"])
        if existing:
            if not replace:
                if existing.get("deck_code") != deck["deck_code"]:
                    raise RuntimeError(
                        f"Deck '{deck['name']}' exists with different code. Use --replace."
                    )
                raise RuntimeError(
                    f"Deck '{deck['name']}' already exists (id={existing['deck_id']})"
                )
            delete_deck(conn, existing["deck_id"])
            conn.commit()

        # Resolve card IDs with exile fallback
        not_found = []
        resolved: list[tuple[DeckCardEntry, str]] = []
        for card in deck["cards"]:
            faction = DECK_NATION_TO_DB.get(card["nation"], card["nation"])
            card_id = find_card_id(conn, faction, card["name"], lang_config.locale_key)
            if card_id is None:
                card_id = find_card_id_by_exile(conn, faction, card["name"], lang_config.locale_key)
            if card_id is None:
                not_found.append(f"{faction} / {card['name']}")
            else:
                resolved.append((card, card_id))

        if not_found:
            lines = "\n".join(f"  - {entry}" for entry in not_found)
            raise RuntimeError(f"Cards not found in collection:\n{lines}")

        # Quantity check
        mismatches: list[tuple[DeckCardEntry, str, int, int]] = []
        for card, card_id in resolved:
            faction = DECK_NATION_TO_DB.get(card["nation"], card["nation"])
            collection_qty = get_card_quantity_by_id(conn, card_id)
            if card["quantity"] != collection_qty:
                mismatches.append((card, card_id, card["quantity"], collection_qty))

        if mismatches and not update:
            lines = "\n".join(
                f"  - {DECK_NATION_TO_DB.get(c['nation'], c['nation'])} / {c['name']}:"
                f" deck={deck_qty}, collection={col_qty}"
                for c, _, deck_qty, col_qty in mismatches
            )
            raise RuntimeError(
                f"Card quantity mismatch:\n{lines}\n"
                "Re-run with --update (-u) to update collection quantities."
            )

        deck_id = insert_deck(conn, deck)
        insert_deck_cards(
            conn,
            deck_id,
            deck["cards"],
            lang_config.locale_key,
            use_exile_fallback=True,
        )
        conn.commit()

        if update and mismatches:
            for _, card_id, deck_qty, _ in mismatches:
                update_card_quantity_by_id(conn, card_id, deck_qty)
            conn.commit()

    logger.info("Deck '%s' added (%d cards)", deck["name"], len(deck["cards"]))


def _select_deck(conn: sqlite3.Connection) -> dict:
    """Interactively select a deck from the database.

    Args:
        conn: SQLite connection instance.

    Returns:
        Selected deck metadata dict.
    """
    decks = fetch_all_decks(conn)
    if not decks:
        raise SystemExit("No decks in database. Run 'kards deck import' first.")

    print("Available decks:")
    for i, deck in enumerate(decks, 1):
        print(f"  {i}. {deck['name']}")

    try:
        choice = int(input("Enter deck number: "))
    except (ValueError, EOFError) as e:
        raise SystemExit(f"Invalid input: expected a deck number (1-{len(decks)})") from e

    if choice < 1 or choice > len(decks):
        raise SystemExit(f"Invalid choice: {choice}. Enter a number from 1 to {len(decks)}")

    return decks[choice - 1]


def remove_deck(db_path: str = DEFAULT_DB_PATH) -> None:
    """Interactively select and delete a deck from the database.

    Enter 0 to delete all decks. Prompts for confirmation before deleting.
    Does not affect the cards table.

    Args:
        db_path: SQLite database path.
    """
    with get_connection(db_path) as conn:
        initialize_schema(conn, db_path)
        decks = fetch_all_decks(conn)
        if not decks:
            raise SystemExit("No decks in database. Run 'kards deck import' first.")

        print("Available decks:")
        print("  0. Delete ALL decks")
        for i, deck in enumerate(decks, 1):
            print(f"  {i}. {deck['name']}")

        try:
            choice = int(input("Enter deck number (0 to delete all): "))
        except (ValueError, EOFError) as e:
            raise SystemExit(f"Invalid input: expected a number (0-{len(decks)})") from e

        if choice < 0 or choice > len(decks):
            raise SystemExit(f"Invalid choice: {choice}. Enter a number from 0 to {len(decks)}")

        if choice == 0:
            print(f"\nAll {len(decks)} deck(s) will be deleted.")
            try:
                confirm = input("Confirm deletion? [y/N]: ").strip().lower()
            except EOFError:
                raise SystemExit("Aborted.")
            if confirm != "y":
                raise SystemExit("Aborted.")
            count = delete_all_decks(conn)
            conn.commit()

        else:
            deck = decks[choice - 1]
            print(f"\nDeck to delete: {deck['name']}")
            try:
                confirm = input("Confirm deletion? [y/N]: ").strip().lower()
            except EOFError:
                raise SystemExit("Aborted.")
            if confirm != "y":
                raise SystemExit("Aborted.")
            delete_deck(conn, deck["deck_id"])
            conn.commit()

    if choice == 0:
        logger.info("All decks deleted (%d).", count)
    else:
        logger.info("Deck '%s' deleted.", deck["name"])


def export_deck(
    fmt: str | None,
    filename: str,
    db_path: str = DEFAULT_DB_PATH,
    *,
    lang: str | None = None,
) -> None:
    """Export a deck to XLSX sheet or JSON file.

    Args:
        fmt: Export format ('json' or None for xlsx).
        filename: Output file path.
        db_path: SQLite database path.
        lang: Active language code (e.g. "en", "ru"). Defaults to English.
    """
    lang_config = get_language_config(lang)
    _emit_locale_warnings(lang_config)

    with get_connection(db_path) as conn:
        initialize_schema(conn, db_path)
        deck_meta = _select_deck(conn)
        deck_cards = fetch_deck_cards(conn, deck_meta["deck_id"])

    if not deck_cards:
        raise SystemExit("Deck has no cards")

    if fmt == "json":
        export_deck_to_json(deck_meta, deck_cards, filename, lang_config)
    else:
        wb = load_workbook(filename)
        add_deck_sheet(
            wb,
            deck_meta,
            deck_cards,
            lang_config.deck_headers,
            lang_config.deck_metadata_labels,
            lang_config.nation_display_names,
            lang_config,
        )
        wb.save(filename)

    logger.info("Deck exported: %s", filename)
