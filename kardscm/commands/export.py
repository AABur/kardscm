"""Export and update-quantity commands."""

from __future__ import annotations

import logging
from pathlib import Path

from openpyxl import load_workbook

from kardscm.commands.utils import _emit_locale_warnings
from kardscm.config import LanguageConfig, get_language_config
from kardscm.constants import DEFAULT_DB_PATH
from kardscm.export import (
    build_collection_headers,
    export_to_json,
    export_to_xlsx,
    translate_card_for_export,
)
from kardscm.helpers import parse_int
from kardscm.storage import (
    fetch_cards,
    get_connection,
    initialize_schema,
    update_quantity,
)

logger = logging.getLogger(__name__)


def _read_xlsx_quantities(
    filename: str,
    lang_config: LanguageConfig,
) -> list[tuple[str, str, int | None]]:
    """Read faction, title, quantity from XLSX file.

    Column names are determined by the active LanguageConfig.
    Returns (faction_display, localized_title, quantity) tuples.

    Args:
        filename: Path to XLSX file.
        lang_config: Active language configuration (drives column headers).

    Returns:
        List of (faction, title, quantity) tuples.

    Raises:
        FileNotFoundError: If file doesn't exist.
        ValueError: If required columns missing.
    """
    if not Path(filename).exists():
        raise FileNotFoundError(f"File not found: {filename}")

    headers_list = lang_config.export_headers
    header_map = {
        headers_list[0]: "faction",
        headers_list[1]: "title",
        headers_list[6]: "quantity",
    }

    wb = load_workbook(filename)
    ws = wb.active
    if not ws:
        raise ValueError("No active worksheet found")

    headers = {
        header_map[cell.value]: col_idx
        for col_idx, cell in enumerate(ws[1], 1)
        if cell.value in header_map
    }
    if len(headers) != len(header_map):
        expected = ", ".join(header_map.keys())
        raise ValueError(f"Missing required columns: {expected}")

    results = []
    for row in ws.iter_rows(min_row=2, values_only=False):
        faction_cell = row[headers["faction"] - 1]
        title_cell = row[headers["title"] - 1]
        qty_cell = row[headers["quantity"] - 1]

        faction = faction_cell.value
        title = title_cell.value
        qty_val = qty_cell.value

        if not faction or not title:
            continue

        qty = parse_int(qty_val)
        results.append((str(faction).strip(), str(title).strip(), qty))

    return results


def export_collection(
    export_format: str,
    filename: str,
    db_path: str = DEFAULT_DB_PATH,
    *,
    lang: str | None = None,
) -> None:
    """Export cards from SQLite to the selected format.

    Args:
        export_format: Export format (json, xlsx).
        filename: Output file path.
        db_path: SQLite database path.
        lang: Active language code (e.g. "en", "ru"). Affects only the XLSX
            export; the JSON export is a raw API-shape dump and is unaffected.
    """
    lang_config = get_language_config(lang)
    _emit_locale_warnings(lang_config)

    with get_connection(db_path) as conn:
        initialize_schema(conn, db_path)
        raw_cards = fetch_cards(conn)

    if not raw_cards:
        raise SystemExit("No cards in database. Run 'kards sync' first.")

    if export_format == "xlsx":
        cards = [translate_card_for_export(card, lang_config) for card in raw_cards]
        export_to_xlsx(
            cards,
            filename,
            build_collection_headers(lang_config),
            lang_config.collection_sheet_name,
        )
    elif export_format == "json":
        export_to_json(raw_cards, filename)
    else:
        msg = f"Unsupported format: {export_format}"
        raise ValueError(msg)

    logger.info("Export completed: %s", filename)


def update_collection(
    filename: str,
    db_path: str = DEFAULT_DB_PATH,
    *,
    lang: str | None = None,
) -> None:
    """Update card quantities from XLSX file.

    Args:
        filename: XLSX file path.
        db_path: SQLite database path.
        lang: Active language code (e.g. "en", "ru"). Defaults to English.
    """
    logger.info("Starting update from file: %s", filename)
    lang_config = get_language_config(lang)
    _emit_locale_warnings(lang_config)

    try:
        updates = _read_xlsx_quantities(filename, lang_config)
    except (FileNotFoundError, ValueError) as e:
        raise SystemExit(f"Failed to read file: {e}") from e

    if not updates:
        logger.warning("No valid entries found in file")
        return

    # Reverse-map localized faction names to API faction names
    reverse_faction = {v: k for k, v in lang_config.faction_names.items()}

    mapped_updates = []
    for faction_display, title, qty in updates:
        faction_api = reverse_faction.get(faction_display, faction_display)
        mapped_updates.append((faction_api, title, qty))

    with get_connection(db_path) as conn:
        initialize_schema(conn, db_path)
        updated_count, not_found = update_quantity(conn, mapped_updates, lang_config.locale_key)

    logger.info(
        "Update completed: %d cards updated, %d not found",
        updated_count,
        len(not_found),
    )

    if not_found:
        for key in not_found:
            logger.warning("Card not found: %s", key)
